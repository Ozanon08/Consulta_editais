import os
import hashlib
import hmac
import logging
import threading
import bcrypt
import smtplib
import base64
from datetime import datetime, timezone
from io import BytesIO

logger = logging.getLogger(__name__)
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

import pandas as pd
import psycopg2
import streamlit as st

# =========================================================
# CONFIG
# =========================================================
st.set_page_config(
    page_title="FGV PMO - Portal de Consulta de Editais",
    page_icon="📘",
    layout="wide",
    initial_sidebar_state="expanded"
)

DATABASE_URL = os.environ.get("DATABASE_URL", "")
SMTP_SERVER = os.getenv("SMTP_SERVER", "")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER", "")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "")
EMAIL_FROM = os.getenv("EMAIL_FROM", SMTP_USER)
LOGO_PATH = os.path.join("assets", "fgv pmo logo.png")


# =========================================================
# ESTILO
# =========================================================
def aplicar_estilo_dark():
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

    /* ── TOKENS ── */
    :root {
        --fgv-navy:      #0b1f3a;
        --fgv-blue:      #1a3f6f;
        --fgv-mid:       #1e5799;
        --fgv-accent:    #2979d4;
        --fgv-bright:    #4d9fff;
        --ink-primary:   #e8edf4;
        --ink-secondary: #94a8c2;
        --ink-muted:     #5a7296;
        --surface-0:     #071628;
        --surface-1:     #0d1e33;
        --surface-2:     #122440;
        --surface-3:     #1a2f52;
        --border:        rgba(41, 121, 212, 0.18);
        --border-strong: rgba(41, 121, 212, 0.35);
        --radius-sm: 6px;
        --radius-md: 10px;
        --radius-lg: 16px;
        --shadow-sm: 0 1px 4px rgba(0,0,0,.3);
        --shadow-md: 0 4px 16px rgba(0,0,0,.4);
        --shadow-lg: 0 8px 32px rgba(0,0,0,.5);
        --font: 'Inter', system-ui, sans-serif;
    }

    /* ── BASE ── */
    .stApp { background: var(--surface-0) !important; color: var(--ink-primary) !important; font-family: var(--font) !important; }
    .block-container { padding: 0 !important; max-width: 100% !important; color: var(--ink-primary) !important; }
    header[data-testid="stHeader"] { background: transparent !important; height: 0 !important; }
    section.main > div { padding-top: 0 !important; }
    div[data-testid="stVerticalBlock"] > div:first-child { margin-top: 0 !important; padding-top: 0 !important; }
    * { font-family: var(--font) !important; }
    p, li { color: var(--ink-primary) !important; }

    /* Esconde botão de colapso da sidebar — todas as variações */
    button[data-testid="collapsedControl"],
    div[data-testid="collapsedControl"],
    [data-testid="stSidebarCollapsedControl"],
    button[data-testid="baseButton-headerNoPadding"],
    div[class*="collapsedControl"],
    span[class*="collapsedControl"],
    [class*="keyboard_double_arrow"],
    .st-emotion-cache-dvne4q,
    .st-emotion-cache-1lna01g { display: none !important; }

    /* Botão que fica flutuando no topo esquerdo da página */
    div[data-testid="stSidebar"] ~ div > button:first-child,
    div[data-testid="stDecoration"] { display: none !important; }

    /* Esconde qualquer botão posicionado absolutamente no canto superior esquerdo */
    section[data-testid="stSidebar"] + div > button { display: none !important; }
    .main > div > button { display: none !important; }

    /* Esconde via atributo de posição — o botão tem position fixed no topo */
    button[style*="top: 0"] { display: none !important; }
    button[style*="top:0"] { display: none !important; }

    /* ── SIDEBAR ── */
    section[data-testid="stSidebar"] {
        background: linear-gradient(170deg, var(--fgv-navy) 0%, var(--fgv-blue) 100%) !important;
        border-right: 1px solid var(--border-strong) !important;
    }
    section[data-testid="stSidebar"] * { color: #fff !important; }
    section[data-testid="stSidebar"] .block-container { padding: 0 0.75rem !important; }
    section[data-testid="stSidebar"] div[data-testid="stVerticalBlock"] > div:first-child { margin-top: -14px !important; }
    .theme-toggle-mini { margin: -6px 4px 2px !important; }
    .sidebar-logo-wrap { display: flex; justify-content: center; padding: 2px 0 10px; }
    .sidebar-logo-img { width: 220px; max-width: 100%; object-fit: contain; }
    section[data-testid="stSidebar"] .stButton > button {
        height: 24px !important; padding: 0 !important; font-size: 12px !important;
        border-radius: var(--radius-sm) !important; background: var(--fgv-accent) !important;
        color: #fff !important; border: none !important; font-weight: 600 !important; box-shadow: none !important;
    }
    section[data-testid="stSidebar"] .stButton > button:hover { background: var(--fgv-mid) !important; }


    /* ── NAV MENU (light) ── */
    section[data-testid="stSidebar"] .stButton > button {
        background: rgba(255,255,255,0.06) !important;
        color: rgba(255,255,255,0.85) !important;
        -webkit-text-fill-color: rgba(255,255,255,0.85) !important;
        border: 1px solid rgba(255,255,255,0.08) !important;
        border-radius: var(--radius-md) !important;
        text-align: center !important;
        justify-content: center !important;
        padding: 9px 14px !important;
        font-size: 0.88rem !important;
        font-weight: 400 !important;
        box-shadow: none !important;
        transform: none !important;
        transition: background .12s, color .12s, border-color .12s !important;
        letter-spacing: 0.01em !important;
        width: 100% !important;
        display: flex !important;
        visibility: visible !important;
        opacity: 1 !important;
    }
    section[data-testid="stSidebar"] .stButton > button:hover {
        background: rgba(255,255,255,0.14) !important;
        color: #fff !important;
        -webkit-text-fill-color: #fff !important;
        border-color: rgba(255,255,255,0.2) !important;
        transform: none !important;
    }

    /* ── NAV ITEM ATIVO ── */
    .nav-ativo {
        background: rgba(255,255,255,0.15) !important;
        border: none !important;
        border-radius: var(--radius-md) !important;
        color: #fff !important;
        font-size: 0.88rem !important;
        font-weight: 700 !important;
        padding: 9px 20px !important;
        margin: 1px 0 !important;
        letter-spacing: 0.01em !important;
        text-align: center !important;
        display: block !important;
        width: 100% !important;
        box-sizing: border-box !important;
        line-height: 1.5 !important;
    }

    /* ── SIDEBAR NAV RADIO ── */
    section[data-testid="stSidebar"] div[role="radiogroup"] { gap: 2px !important; }
    section[data-testid="stSidebar"] label[data-testid="stWidgetLabel"] { display: none !important; }
    section[data-testid="stSidebar"] div[data-testid="stRadio"] > div { gap: 2px !important; }
    section[data-testid="stSidebar"] div[data-testid="stRadio"] label {
        display: flex !important; align-items: center !important;
        padding: 8px 12px !important; border-radius: var(--radius-md) !important;
        cursor: pointer !important; transition: background .15s !important;
        font-size: 0.9rem !important; font-weight: 500 !important;
        color: rgba(255,255,255,0.82) !important;
        -webkit-text-fill-color: rgba(255,255,255,0.82) !important;
        margin: 0 !important; width: 100% !important;
    }
    section[data-testid="stSidebar"] div[data-testid="stRadio"] label:hover {
        background: rgba(255,255,255,0.1) !important;
        color: #fff !important; -webkit-text-fill-color: #fff !important;
    }
    section[data-testid="stSidebar"] div[data-testid="stRadio"] label:has(input:checked) {
        background: rgba(41,121,212,0.35) !important;
        color: #fff !important; -webkit-text-fill-color: #fff !important;
        font-weight: 600 !important;
        border-left: 3px solid rgba(77,159,255,0.9) !important;
    }
    section[data-testid="stSidebar"] div[data-testid="stRadio"] input[type="radio"] {
        display: none !important;
    }
    section[data-testid="stSidebar"] div[data-testid="stRadio"] div[data-testid="stMarkdownContainer"] {
        display: none !important;
    }


    /* ── NAV MENU (dark) ── */
    section[data-testid="stSidebar"] .stButton > button {
        background: transparent !important;
        color: rgba(255,255,255,0.75) !important;
        -webkit-text-fill-color: rgba(255,255,255,0.75) !important;
        border: none !important;
        border-radius: var(--radius-md) !important;
        text-align: center !important;
        justify-content: center !important;
        padding: 9px 14px !important;
        font-size: 0.88rem !important;
        font-weight: 400 !important;
        box-shadow: none !important;
        transform: none !important;
        transition: background .12s, color .12s !important;
        letter-spacing: 0.01em !important;
        width: 100% !important;
        display: flex !important;
        visibility: visible !important;
        opacity: 1 !important;
    }
    section[data-testid="stSidebar"] .stButton > button:hover {
        background: rgba(255,255,255,0.1) !important;
        color: #fff !important;
        -webkit-text-fill-color: #fff !important;
        transform: none !important;
    }

    /* ── NAV ITEM ATIVO ── */
    .nav-ativo {
        background: rgba(255,255,255,0.15) !important;
        border: none !important;
        border-radius: var(--radius-md) !important;
        color: #fff !important;
        font-size: 0.88rem !important;
        font-weight: 700 !important;
        padding: 9px 20px !important;
        margin: 1px 0 !important;
        letter-spacing: 0.01em !important;
        text-align: center !important;
        display: block !important;
        width: 100% !important;
        box-sizing: border-box !important;
        line-height: 1.5 !important;
    }

    /* ── SIDEBAR NAV RADIO (dark) ── */
    section[data-testid="stSidebar"] div[data-testid="stRadio"] > div { gap: 2px !important; }
    section[data-testid="stSidebar"] div[data-testid="stRadio"] label {
        display: flex !important; align-items: center !important;
        padding: 8px 12px !important; border-radius: var(--radius-md) !important;
        cursor: pointer !important; transition: background .15s !important;
        font-size: 0.9rem !important; font-weight: 500 !important;
        color: rgba(255,255,255,0.78) !important;
        -webkit-text-fill-color: rgba(255,255,255,0.78) !important;
        margin: 0 !important; width: 100% !important;
    }
    section[data-testid="stSidebar"] div[data-testid="stRadio"] label:hover {
        background: rgba(255,255,255,0.08) !important;
        color: #fff !important; -webkit-text-fill-color: #fff !important;
    }
    section[data-testid="stSidebar"] div[data-testid="stRadio"] label:has(input:checked) {
        background: rgba(41,121,212,0.28) !important;
        color: #fff !important; -webkit-text-fill-color: #fff !important;
        font-weight: 600 !important; border-left: 3px solid var(--fgv-bright) !important;
    }
    section[data-testid="stSidebar"] div[data-testid="stRadio"] input[type="radio"] { display: none !important; }

        /* ── HEADER ── */
    .header-full-width {
        background: linear-gradient(100deg, var(--fgv-navy) 0%, var(--fgv-blue) 50%, var(--fgv-mid) 100%);
        padding: 28px 36px 24px; color: #fff; min-height: 128px;
        border-bottom: 1px solid var(--border-strong);
    }
    .header-inner { display: flex; align-items: center; justify-content: space-between; gap: 24px; }
    .header-text-block { flex: 1; min-width: 0; }
    .header-logo-block { flex-shrink: 0; display: flex; align-items: center; }
    .header-logo-full { height: 64px; max-width: 240px; object-fit: contain; opacity: .95; }
    .header-title { font-size: 1.75rem; font-weight: 700; color: #fff; margin-bottom: 6px; letter-spacing: -0.02em; line-height: 1.2; }
    .header-subtitle { font-size: 0.9rem; color: rgba(255,255,255,.75); margin-bottom: 12px; }
    .header-profile {
        display: inline-flex; align-items: center; gap: 6px;
        background: rgba(255,255,255,.1); border: 1px solid rgba(255,255,255,.15);
        padding: 5px 12px; border-radius: 999px; font-size: 0.85rem; color: #fff;
    }

    /* ── SECTION CARDS ── */
    .section-card {
        background: var(--surface-1); border: 1px solid var(--border);
        border-radius: var(--radius-lg); padding: 20px;
        margin: 0 16px 14px; box-shadow: var(--shadow-sm);
    }
    .metric-card {
        background: var(--surface-2); border: 1px solid var(--border);
        border-radius: var(--radius-md); padding: 16px; box-shadow: var(--shadow-sm);
    }
    .metric-title { color: var(--ink-secondary) !important; font-size: 0.8rem; font-weight: 600; text-transform: uppercase; letter-spacing: .04em; }
    .metric-value { color: var(--fgv-bright) !important; font-weight: 800; font-size: 1.7rem; line-height: 1.1; }
    .metric-sub { color: var(--ink-muted) !important; font-size: 0.8rem; margin-top: 2px; }

    /* ── INPUTS ── */
    .stTextInput input, .stNumberInput input, .stTextArea textarea,
    input[type="text"], input[type="number"], textarea {
        background: var(--surface-2) !important; color: var(--ink-primary) !important;
        -webkit-text-fill-color: var(--ink-primary) !important;
        border: 1px solid var(--border-strong) !important; border-radius: var(--radius-md) !important;
        font-family: var(--font) !important;
    }
    .stTextInput input:focus, .stNumberInput input:focus, .stTextArea textarea:focus {
        border-color: var(--fgv-accent) !important; box-shadow: 0 0 0 2px rgba(41,121,212,.2) !important;
    }
    .stTextInput input::placeholder, .stTextArea textarea::placeholder { color: var(--ink-muted) !important; opacity: 1 !important; }
    .stSelectbox > div > div, div[data-baseweb="select"] > div {
        background: var(--surface-2) !important; border: 1px solid var(--border-strong) !important;
        border-radius: var(--radius-md) !important;
    }
    div[data-baseweb="select"] * { color: var(--ink-primary) !important; -webkit-text-fill-color: var(--ink-primary) !important; }
    div[data-baseweb="select"] svg { fill: var(--fgv-bright) !important; }
    div[data-baseweb="menu"] {
        background: var(--surface-1) !important; border: 1px solid var(--border-strong) !important;
        border-radius: var(--radius-md) !important; box-shadow: var(--shadow-lg) !important;
    }
    div[data-baseweb="menu"] * { color: var(--ink-primary) !important; -webkit-text-fill-color: var(--ink-primary) !important; }
    div[data-baseweb="menu"] li:hover { background: var(--surface-3) !important; }
    label { color: var(--ink-secondary) !important; font-size: 0.85rem !important; font-weight: 500 !important; }
    .stNumberInput button { background: var(--surface-2) !important; color: var(--ink-primary) !important; border: 1px solid var(--border-strong) !important; }

    /* ── BUTTONS ── */
    .stButton > button, .stDownloadButton > button, .stFormSubmitButton > button {
        background: var(--fgv-accent) !important; color: #fff !important;
        -webkit-text-fill-color: #fff !important; border: none !important;
        border-radius: var(--radius-md) !important; font-weight: 600 !important;
        font-family: var(--font) !important; transition: background .15s, transform .1s !important;
        box-shadow: 0 2px 8px rgba(41,121,212,.3) !important;
    }
    .stButton > button:hover, .stDownloadButton > button:hover { background: var(--fgv-mid) !important; transform: translateY(-1px) !important; }
    .stButton > button:active { transform: translateY(0) !important; }
    .stButton > button:disabled { background: var(--surface-3) !important; color: var(--ink-muted) !important; -webkit-text-fill-color: var(--ink-muted) !important; box-shadow: none !important; transform: none !important; }


    /* ── SELECTBOX MODERNO (dark) ── */
    div[data-baseweb="select"] {
        border-radius: var(--radius-md) !important;
    }
    div[data-baseweb="select"] > div {
        background: var(--surface-2) !important;
        border: 1.5px solid var(--border-strong) !important;
        border-radius: var(--radius-md) !important;
        min-height: 42px !important;
        transition: border-color .15s, box-shadow .15s !important;
    }
    div[data-baseweb="select"] > div:hover {
        border-color: var(--fgv-accent) !important;
    }
    div[data-baseweb="select"] > div:focus-within {
        border-color: var(--fgv-accent) !important;
        box-shadow: 0 0 0 3px rgba(41,121,212,.2) !important;
    }
    div[data-baseweb="select"] [data-testid="stSelectboxVirtualDropdown"],
    div[data-baseweb="popover"] {
        background: var(--surface-1) !important;
        border: 1px solid var(--border-strong) !important;
        border-radius: var(--radius-md) !important;
        box-shadow: 0 8px 24px rgba(0,0,0,.5) !important;
        padding: 4px !important;
    }
    div[data-baseweb="option"] {
        border-radius: var(--radius-sm) !important;
        margin: 1px 4px !important;
        padding: 8px 12px !important;
        transition: background .1s !important;
    }
    div[data-baseweb="option"]:hover {
        background: var(--surface-3) !important;
    }
    div[data-baseweb="option"][aria-selected="true"] {
        background: rgba(41,121,212,.25) !important;
    }
    /* ── TABLE ── */
    div[data-testid="stDataFrame"] { background: var(--surface-1) !important; border: 1px solid var(--border) !important; border-radius: var(--radius-md) !important; }
    div[data-testid="stDataFrame"] div[role="grid"] { background: var(--surface-1) !important; }
    div[data-testid="stDataFrame"] div[role="row"] { background: var(--surface-1) !important; }
    div[data-testid="stDataFrame"] div[role="row"]:nth-child(even) { background: var(--surface-2) !important; }
    div[data-testid="stDataFrame"] div[role="row"]:hover { background: var(--surface-3) !important; }
    div[data-testid="stDataFrame"] div[role="gridcell"] { color: var(--ink-primary) !important; border-bottom: 1px solid var(--border) !important; }
    div[data-testid="stDataFrame"] div[role="columnheader"] { background: var(--surface-3) !important; color: var(--fgv-bright) !important; font-weight: 600 !important; font-size: 0.8rem !important; text-transform: uppercase !important; letter-spacing: .04em !important; border-bottom: 1px solid var(--border-strong) !important; }

    /* ── LOGIN ── */
    .login-card { background: var(--surface-1) !important; border: 1px solid var(--border-strong) !important; border-radius: var(--radius-lg) !important; box-shadow: var(--shadow-lg) !important; padding: 10px !important; }
    .login-title { color: var(--fgv-bright) !important; font-size: 1.6rem !important; font-weight: 700 !important; letter-spacing: -0.02em !important; }
    .login-subtitle, .small-muted { color: var(--ink-secondary) !important; }

    /* ── ALERTS ── */
    div[data-testid="stAlert"] { border-radius: var(--radius-md) !important; border-left-width: 3px !important; }

    /* ── EXPANDER ── */
    div[data-testid="stExpander"] { background: var(--surface-2) !important; border: 1px solid var(--border) !important; border-radius: var(--radius-md) !important; }

    /* ── METRICS ── */
    div[data-testid="stMetric"] { background: var(--surface-2) !important; border: 1px solid var(--border) !important; border-radius: var(--radius-md) !important; padding: 12px 16px !important; }
    div[data-testid="stMetric"] label { color: var(--ink-secondary) !important; }
    div[data-testid="stMetric"] [data-testid="stMetricValue"] { color: var(--fgv-bright) !important; }

    </style>
    """, unsafe_allow_html=True)


def aplicar_estilo_light():
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

    /* ── TOKENS ── */
    :root {
        --fgv-navy:      #0b1f3a;
        --fgv-blue:      #1a3f6f;
        --fgv-mid:       #1e5799;
        --fgv-accent:    #1d6fc4;
        --fgv-bright:    #1a5fa8;
        --ink-primary:   #0d1b2e;
        --ink-secondary: #3d5575;
        --ink-muted:     #7a97b8;
        --surface-0:     #f0f4f9;
        --surface-1:     #ffffff;
        --surface-2:     #e8eef6;
        --surface-3:     #d8e5f2;
        --border:        rgba(26, 63, 111, 0.12);
        --border-strong: rgba(26, 63, 111, 0.22);
        --radius-sm: 6px;
        --radius-md: 10px;
        --radius-lg: 16px;
        --shadow-sm: 0 1px 4px rgba(11,31,58,.06);
        --shadow-md: 0 4px 16px rgba(11,31,58,.09);
        --shadow-lg: 0 8px 32px rgba(11,31,58,.12);
        --font: 'Inter', system-ui, sans-serif;
    }

    /* ── BASE ── */
    .stApp { background: var(--surface-0) !important; color: var(--ink-primary) !important; font-family: var(--font) !important; }
    .block-container { padding: 0 !important; max-width: 100% !important; color: var(--ink-primary) !important; }
    header[data-testid="stHeader"] { background: transparent !important; height: 0 !important; }
    section.main > div { padding-top: 0 !important; }
    div[data-testid="stVerticalBlock"] > div:first-child { margin-top: 0 !important; padding-top: 0 !important; }
    * { font-family: var(--font) !important; }
    p, li { color: var(--ink-primary) !important; }

    /* Esconde botão de colapso da sidebar — todas as variações */
    button[data-testid="collapsedControl"],
    div[data-testid="collapsedControl"],
    [data-testid="stSidebarCollapsedControl"],
    button[data-testid="baseButton-headerNoPadding"],
    div[class*="collapsedControl"],
    span[class*="collapsedControl"],
    [class*="keyboard_double_arrow"],
    .st-emotion-cache-dvne4q,
    .st-emotion-cache-1lna01g { display: none !important; }

    /* Botão que fica flutuando no topo esquerdo da página */
    div[data-testid="stSidebar"] ~ div > button:first-child,
    div[data-testid="stDecoration"] { display: none !important; }

    /* Esconde qualquer botão posicionado absolutamente no canto superior esquerdo */
    section[data-testid="stSidebar"] + div > button { display: none !important; }
    .main > div > button { display: none !important; }

    /* Esconde via atributo de posição — o botão tem position fixed no topo */
    button[style*="top: 0"] { display: none !important; }
    button[style*="top:0"] { display: none !important; }

    /* ── SIDEBAR ── */
    section[data-testid="stSidebar"] {
        background: linear-gradient(170deg, var(--fgv-navy) 0%, var(--fgv-blue) 100%) !important;
        border-right: 1px solid var(--border-strong) !important;
    }
    section[data-testid="stSidebar"] * { color: #fff !important; }
    section[data-testid="stSidebar"] .block-container { padding: 0 0.75rem !important; }
    section[data-testid="stSidebar"] div[data-testid="stVerticalBlock"] > div:first-child { margin-top: -14px !important; }
    .theme-toggle-mini { margin: -6px 4px 2px !important; }
    .sidebar-logo-wrap { display: flex; justify-content: center; padding: 2px 0 10px; }
    .sidebar-logo-img { width: 220px; max-width: 100%; object-fit: contain; filter: brightness(1.3); }
    section[data-testid="stSidebar"] .stButton > button {
        height: 24px !important; padding: 0 !important; font-size: 12px !important;
        border-radius: var(--radius-sm) !important; background: var(--fgv-accent) !important;
        color: #fff !important; -webkit-text-fill-color: #fff !important;
        border: none !important; font-weight: 600 !important; box-shadow: none !important;
    }
    section[data-testid="stSidebar"] .stButton > button:hover { background: var(--fgv-mid) !important; }

    /* ── HEADER ── */
    .header-full-width {
        background: linear-gradient(100deg, var(--fgv-navy) 0%, var(--fgv-blue) 50%, var(--fgv-mid) 100%);
        padding: 28px 36px 24px; color: #fff; min-height: 128px;
        border-bottom: 2px solid rgba(255,255,255,.08);
    }
    .header-inner { display: flex; align-items: center; justify-content: space-between; gap: 24px; }
    .header-text-block { flex: 1; min-width: 0; }
    .header-logo-block { flex-shrink: 0; display: flex; align-items: center; }
    .header-logo-full { height: 64px; max-width: 240px; object-fit: contain; opacity: .95; }
    .header-title { font-size: 1.75rem; font-weight: 700; color: #fff; margin-bottom: 6px; letter-spacing: -0.02em; line-height: 1.2; }
    .header-subtitle { font-size: 0.9rem; color: rgba(255,255,255,.78); margin-bottom: 12px; }
    .header-profile {
        display: inline-flex; align-items: center; gap: 6px;
        background: rgba(255,255,255,.12); border: 1px solid rgba(255,255,255,.18);
        padding: 5px 12px; border-radius: 999px; font-size: 0.85rem; color: #fff;
    }

    /* ── SECTION CARDS ── */
    .section-card {
        background: var(--surface-1); border: 1px solid var(--border);
        border-radius: var(--radius-lg); padding: 20px;
        margin: 0 16px 14px; box-shadow: var(--shadow-sm);
    }
    .metric-card {
        background: var(--surface-1); border: 1px solid var(--border);
        border-radius: var(--radius-md); padding: 16px; box-shadow: var(--shadow-sm);
    }
    .metric-title { color: var(--ink-secondary) !important; font-size: 0.8rem; font-weight: 600; text-transform: uppercase; letter-spacing: .04em; }
    .metric-value { color: var(--fgv-blue) !important; font-weight: 800; font-size: 1.7rem; line-height: 1.1; }
    .metric-sub { color: var(--ink-muted) !important; font-size: 0.8rem; margin-top: 2px; }

    /* ── INPUTS ── */
    .stTextInput input, .stNumberInput input, .stTextArea textarea,
    input[type="text"], input[type="number"], textarea {
        background: #f5f8fd !important; color: var(--ink-primary) !important;
        -webkit-text-fill-color: var(--ink-primary) !important;
        border: 1px solid var(--border-strong) !important; border-radius: var(--radius-md) !important;
        font-family: var(--font) !important;
    }
    .stTextInput input:focus, .stNumberInput input:focus, .stTextArea textarea:focus {
        border-color: var(--fgv-accent) !important; box-shadow: 0 0 0 2px rgba(29,111,196,.15) !important;
    }
    .stTextInput input::placeholder, .stTextArea textarea::placeholder { color: var(--ink-muted) !important; opacity: 1 !important; }
    .stSelectbox > div > div, div[data-baseweb="select"] > div {
        background: #f5f8fd !important; border: 1px solid var(--border-strong) !important;
        border-radius: var(--radius-md) !important;
    }
    div[data-baseweb="select"] * { color: var(--ink-primary) !important; -webkit-text-fill-color: var(--ink-primary) !important; }
    div[data-baseweb="select"] svg { fill: var(--fgv-blue) !important; }
    div[data-baseweb="menu"], div[data-baseweb="popover"] div[data-baseweb="menu"] {
        background: #fff !important; border: 1px solid var(--border-strong) !important;
        border-radius: var(--radius-md) !important; box-shadow: var(--shadow-lg) !important;
    }
    div[data-baseweb="menu"] *, div[data-baseweb="popover"] div[data-baseweb="menu"] * {
        color: var(--ink-primary) !important; -webkit-text-fill-color: var(--ink-primary) !important;
        background: transparent !important;
    }
    div[data-baseweb="menu"] li:hover, div[data-baseweb="popover"] li:hover { background: var(--surface-2) !important; }
    label { color: var(--ink-secondary) !important; font-size: 0.85rem !important; font-weight: 500 !important; }
    .stMarkdown p, .stMarkdown li, .stMarkdown span { color: var(--ink-primary) !important; }
    .stNumberInput button { background: #f5f8fd !important; color: var(--ink-primary) !important; border: 1px solid var(--border-strong) !important; }

    /* ── BUTTONS ── */
    .stButton > button, .stDownloadButton > button, .stFormSubmitButton > button {
        background: var(--fgv-accent) !important; color: #fff !important;
        -webkit-text-fill-color: #fff !important; border: none !important;
        border-radius: var(--radius-md) !important; font-weight: 600 !important;
        font-family: var(--font) !important; transition: background .15s, transform .1s !important;
        box-shadow: 0 2px 8px rgba(29,111,196,.25) !important;
    }
    .stButton > button:hover, .stDownloadButton > button:hover { background: var(--fgv-blue) !important; transform: translateY(-1px) !important; }
    .stButton > button:active { transform: translateY(0) !important; }
    .stButton > button:disabled { background: var(--surface-2) !important; color: var(--ink-muted) !important; -webkit-text-fill-color: var(--ink-muted) !important; box-shadow: none !important; transform: none !important; }


    /* ── SELECTBOX MODERNO (light) ── */
    div[data-baseweb="select"] {
        border-radius: var(--radius-md) !important;
    }
    div[data-baseweb="select"] > div {
        background: #f5f8fd !important;
        border: 1.5px solid var(--border-strong) !important;
        border-radius: var(--radius-md) !important;
        min-height: 42px !important;
        transition: border-color .15s, box-shadow .15s !important;
    }
    div[data-baseweb="select"] > div:hover {
        border-color: var(--fgv-accent) !important;
    }
    div[data-baseweb="select"] > div:focus-within {
        border-color: var(--fgv-accent) !important;
        box-shadow: 0 0 0 3px rgba(29,111,196,.15) !important;
    }
    div[data-baseweb="popover"],
    div[data-baseweb="select"] [data-testid="stSelectboxVirtualDropdown"] {
        background: #fff !important;
        border: 1px solid var(--border-strong) !important;
        border-radius: var(--radius-md) !important;
        box-shadow: 0 8px 24px rgba(11,31,58,.12) !important;
        padding: 4px !important;
    }
    div[data-baseweb="option"] {
        border-radius: var(--radius-sm) !important;
        margin: 1px 4px !important;
        padding: 8px 12px !important;
        transition: background .1s !important;
        color: var(--ink-primary) !important;
    }
    div[data-baseweb="option"]:hover {
        background: var(--surface-2) !important;
    }
    div[data-baseweb="option"][aria-selected="true"] {
        background: var(--surface-3) !important;
        color: var(--fgv-blue) !important;
        font-weight: 600 !important;
    }
    /* ── TABLE ── */
    div[data-testid="stDataFrame"] { background: #fff !important; border: 1px solid var(--border) !important; border-radius: var(--radius-md) !important; }
    div[data-testid="stDataFrame"] div[role="grid"] { background: #fff !important; }
    div[data-testid="stDataFrame"] div[role="row"] { background: #fff !important; }
    div[data-testid="stDataFrame"] div[role="row"]:nth-child(even) { background: var(--surface-2) !important; }
    div[data-testid="stDataFrame"] div[role="row"]:hover { background: var(--surface-3) !important; }
    div[data-testid="stDataFrame"] div[role="gridcell"] { color: var(--ink-primary) !important; -webkit-text-fill-color: var(--ink-primary) !important; border-bottom: 1px solid var(--border) !important; background: transparent !important; }
    div[data-testid="stDataFrame"] div[role="gridcell"] * { color: var(--ink-primary) !important; -webkit-text-fill-color: var(--ink-primary) !important; }
    div[data-testid="stDataFrame"] div[role="columnheader"] { background: var(--surface-2) !important; color: var(--fgv-blue) !important; -webkit-text-fill-color: var(--fgv-blue) !important; font-weight: 700 !important; font-size: 0.8rem !important; text-transform: uppercase !important; letter-spacing: .04em !important; border-bottom: 2px solid var(--border-strong) !important; }
    div[data-testid="stDataFrame"] div[role="columnheader"] * { color: var(--fgv-blue) !important; -webkit-text-fill-color: var(--fgv-blue) !important; }

    /* ── LOGIN ── */
    .login-card { background: #fff !important; border: 1px solid var(--border) !important; border-radius: var(--radius-lg) !important; box-shadow: var(--shadow-lg) !important; padding: 10px !important; }
    .login-title { color: var(--fgv-blue) !important; font-size: 1.6rem !important; font-weight: 700 !important; letter-spacing: -0.02em !important; }
    .login-subtitle, .small-muted, .login-footer { color: var(--ink-secondary) !important; }
    .login-card .stTextInput input { background: #f5f8fd !important; border: 1px solid var(--border-strong) !important; }

    /* ── ALERTS ── */
    div[data-testid="stAlert"] { border-radius: var(--radius-md) !important; border-left-width: 3px !important; }

    /* ── EXPANDER ── */
    div[data-testid="stExpander"] { background: var(--surface-1) !important; border: 1px solid var(--border) !important; border-radius: var(--radius-md) !important; }

    /* ── METRICS ── */
    div[data-testid="stMetric"] { background: var(--surface-1) !important; border: 1px solid var(--border) !important; border-radius: var(--radius-md) !important; padding: 12px 16px !important; }
    div[data-testid="stMetric"] label { color: var(--ink-secondary) !important; }
    div[data-testid="stMetric"] [data-testid="stMetricValue"] { color: var(--fgv-blue) !important; }

    </style>
    """, unsafe_allow_html=True)


def aplicar_estilo(modo="light"):
    if modo == "dark":
        aplicar_estilo_dark()
    else:
        aplicar_estilo_light()


# =========================================================
# BANCO / UTIL
# =========================================================
def get_conn():
    return psycopg2.connect(DATABASE_URL)


def hash_senha(senha: str) -> str:
    """Retorna o hash bcrypt da senha."""
    return bcrypt.hashpw(senha.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verificar_senha(senha: str, hash_armazenado: str) -> bool:
    """Verifica senha contra hash bcrypt ou SHA-256 legado."""
    try:
        return bcrypt.checkpw(senha.encode("utf-8"), hash_armazenado.encode("utf-8"))
    except Exception:
        calculado = hashlib.sha256(senha.encode("utf-8")).hexdigest()
        return hmac.compare_digest(calculado, hash_armazenado or "")


# ── Rate limiting em memória ──────────────────────────────
_DUMMY_BCRYPT_HASH = "$2b$12$KIXdR5v2FJjBi3vkHxBnL.aBGz8zVkxRZl1GQoqsJZJw5c5gVkIUC"
_login_lock = threading.Lock()
_login_attempts: dict = {}
_MAX_TENTATIVAS = 5
_JANELA_SEGUNDOS = 300


def _checar_e_registrar_tentativa(username: str, sucesso: bool):
    import time
    agora = time.time()
    key = username.strip().upper()
    with _login_lock:
        tentativas = _login_attempts.get(key, [])
        tentativas = [t for t in tentativas if agora - t < _JANELA_SEGUNDOS]
        if sucesso:
            _login_attempts[key] = []
            return False, 0
        tentativas.append(agora)
        _login_attempts[key] = tentativas
        if len(tentativas) >= _MAX_TENTATIVAS:
            restante = int(_JANELA_SEGUNDOS - (agora - tentativas[0]))
            return True, max(0, restante)
        return False, 0


def _dummy_bcrypt(senha: str) -> None:
    try:
        bcrypt.checkpw(senha.encode("utf-8"), _DUMMY_BCRYPT_HASH.encode("utf-8"))
    except Exception:
        pass


def agora_str() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")


def init_db():
    """Garante que o usuário ADMIN existe. Tabelas já criadas via schema_supabase.sql."""
    if not DATABASE_URL:
        st.error("Variável DATABASE_URL não configurada. Defina-a nas configurações do Streamlit Cloud.")
        st.stop()
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM usuarios WHERE username = 'ADMIN'")
    exists = cur.fetchone()[0]
    if exists == 0:
        senha_inicial = os.environ.get("ADMIN_INITIAL_PASSWORD", "")
        if not senha_inicial:
            conn.close()
            st.error("Defina ADMIN_INITIAL_PASSWORD nas variáveis de ambiente do Streamlit Cloud.")
            st.stop()
        cur.execute("""
            INSERT INTO usuarios (username, email, senha_hash, perfil, ativo, criado_em, atualizado_em)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, ("ADMIN", "", hash_senha(senha_inicial), "ADMIN", 1, agora_str(), agora_str()))
    conn.commit()
    conn.close()


@st.cache_data(ttl=300, show_spinner=False)
def carregar_view():
    conn = get_conn()
    try:
        df = pd.read_sql_query("SELECT * FROM vw_consulta_editais", conn)
    except Exception:
        df = pd.DataFrame()
    finally:
        conn.close()
    return df


def autenticar(username: str, senha: str):
    # Rate limiting antes de qualquer consulta ao banco
    bloqueado, restante = _checar_e_registrar_tentativa(username, sucesso=False)
    if bloqueado:
        raise PermissionError(f"Muitas tentativas incorretas. Aguarde {restante} segundos.")

    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT username, perfil, ativo, email, senha_hash
        FROM usuarios
        WHERE UPPER(username) = UPPER(%s)
    """, (username,))
    row = cur.fetchone()

    if not row:
        # Usuário não existe — dummy bcrypt para nivelar tempo (evita username enumeration)
        _dummy_bcrypt(senha)
        conn.close()
        return None

    if row[2] == 1 and verificar_senha(senha, row[4]):
        _checar_e_registrar_tentativa(username, sucesso=True)
        # Upgrade automático SHA-256 → bcrypt
        if not row[4].startswith("$2b$") and not row[4].startswith("$2a$"):
            novo_hash = hash_senha(senha)
            cur.execute("UPDATE usuarios SET senha_hash = %s WHERE UPPER(username) = UPPER(%s)",
                       (novo_hash, row[0]))
            conn.commit()
        conn.close()
        return {"username": row[0], "perfil": row[1], "email": row[3] if len(row) > 3 else ""}

    conn.close()
    return None


def listar_usuarios():
    conn = get_conn()
    df = pd.read_sql_query("""
        SELECT id, username, COALESCE(email, '') AS email, perfil, ativo, criado_em
        FROM usuarios
        ORDER BY username
    """, conn)
    conn.close()
    return df


def validar_senha(senha: str) -> tuple[bool, str]:
    if len(senha) < 8:
        return False, "A senha deve ter pelo menos 8 caracteres."
    if not any(c.isupper() for c in senha):
        return False, "A senha deve conter pelo menos uma letra maiúscula."
    if not any(c.islower() for c in senha):
        return False, "A senha deve conter pelo menos uma letra minúscula."
    if not any(c.isdigit() for c in senha):
        return False, "A senha deve conter pelo menos um número."
    return True, ""


def criar_usuario(username: str, email: str, senha: str, perfil: str):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO usuarios (username, email, senha_hash, perfil, ativo, criado_em, atualizado_em)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
    """, (username.strip(), email.strip(), hash_senha(senha), perfil, 1, agora_str(), agora_str()))
    conn.commit()
    conn.close()


def alterar_status_usuario(user_id: int, ativo: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("UPDATE usuarios SET ativo = %s, atualizado_em = %s WHERE id = %s", (ativo, agora_str(), user_id))
    conn.commit()
    conn.close()


def excluir_usuario(user_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM usuarios WHERE id = %s", (user_id,))
    conn.commit()
    conn.close()


def alterar_senha_usuario(username: str, senha_atual: str, nova_senha: str):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT senha_hash FROM usuarios WHERE UPPER(username) = UPPER(%s)", (username,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return False, "Usuário não encontrado."
    if row[0] != hash_senha(senha_atual):
        conn.close()
        return False, "Senha atual incorreta."
    cur.execute(
        "UPDATE usuarios SET senha_hash = %s, atualizado_em = %s WHERE UPPER(username) = UPPER(%s)",
        (hash_senha(nova_senha), agora_str(), username)
    )
    conn.commit()
    conn.close()
    return True, "Senha alterada com sucesso."


def inserir_solicitacao(tema: str, descricao: str, solicitante: str, perfil: str):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO solicitacoes_tema
            (tema_solicitado, descricao, solicitante, perfil_solicitante, status, data_solicitacao)
        VALUES (%s, %s, %s, %s, 'PENDENTE', %s)
        RETURNING id
    """, (tema.strip(), descricao.strip(), solicitante, perfil, agora_str()))
    solicitacao_id = cur.fetchone()[0]
    conn.commit()
    conn.close()
    return solicitacao_id


@st.cache_data(ttl=60, show_spinner=False)
def listar_solicitacoes():
    conn = get_conn()
    df = pd.read_sql_query("""
        SELECT id, tema_solicitado, descricao, solicitante, perfil_solicitante, status, data_solicitacao
        FROM solicitacoes_tema
        ORDER BY id DESC
    """, conn)
    conn.close()
    return df


def atualizar_status_solicitacao(solicitacao_id: int, novo_status: str):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("UPDATE solicitacoes_tema SET status = %s WHERE id = %s", (novo_status, solicitacao_id))
    conn.commit()
    conn.close()


def obter_solicitacao_por_id(solicitacao_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, tema_solicitado, descricao, solicitante, perfil_solicitante, status, data_solicitacao
        FROM solicitacoes_tema
        WHERE id = %s
    """, (solicitacao_id,))
    row = cur.fetchone()
    conn.close()
    return row


def buscar_emails_admins():
    conn = get_conn()
    df = pd.read_sql_query("""
        SELECT email FROM usuarios
        WHERE perfil = 'ADMIN' AND ativo = 1
          AND email IS NOT NULL AND TRIM(email) <> ''
    """, conn)
    conn.close()
    return df["email"].tolist()


def buscar_email_usuario(username: str):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT email FROM usuarios WHERE UPPER(username) = UPPER(%s)", (username,))
    row = cur.fetchone()
    conn.close()
    return row[0] if row and row[0] else ""


def normalizar_servicos(value) -> list:
    if not value or str(value).strip() in ("", "nan", "None", "-"):
        return []
    text = str(value).strip().rstrip(".")
    separators = [",", ";", "|", " / "]
    parts = [text]
    for sep in separators:
        new_parts = []
        for item in parts:
            new_parts.extend(item.split(sep))
        parts = new_parts
    parts = [p.strip(" .;") for p in parts if str(p).strip(" .;")]
    seen = set()
    result = []
    for part in parts:
        key = part.casefold()
        if key not in seen:
            seen.add(key)
            result.append(part)
    return result


def processar_upload_planilha(arquivo):
    import pandas as pd

    COLUMN_MAP = {
        "Tema": "tema", "Subtema": "subtema", "Serviços": "servicos",
        "País": "pais", "Estado": "estado", "Município": "municipio",
        "Nome Edital": "nome_edital", "Descrição": "descricao", "Esforço": "esforco",
        "Unidade": "unidade", "Prazo (meses)": "prazo_meses",
        "Tipo de Edital": "tipo_edital", "Código Planilha": "codigo_planilha",
        "Fonte de Dados": "fonte_dado", "OBS": "observacao",
        "Custo de Execução": "custo_execucao", "Data edital (mês/ano)": "data_edital",
        "Min": "valor_min", "Máx": "valor_max",
        "Metodo de Calculo": "metodo_calculo",
        "Método de Cálculo": "metodo_calculo",
    }

    df = pd.read_excel(arquivo, sheet_name="Base")
    df = df.rename(columns=COLUMN_MAP)
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].astype(str).str.strip().replace({"nan": None, "None": None, "": None})
    for col in ["prazo_meses", "custo_execucao", "valor_min", "valor_max"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    if "data_edital" in df.columns:
        dt = pd.to_datetime(df["data_edital"], errors="coerce", dayfirst=True)
        mask = dt.notna()
        df.loc[mask, "data_edital"] = dt.loc[mask].dt.strftime("%Y-%m")

    conn = get_conn()
    conn.autocommit = False
    cur = conn.cursor()

    # Limpa apenas os editais (preserva usuários e solicitações)
    cur.execute("DELETE FROM edital_servico")
    cur.execute("DELETE FROM edital")
    cur.execute("DELETE FROM servico")
    cur.execute("DELETE FROM fonte_dado")
    cur.execute("DELETE FROM unidade")
    cur.execute("DELETE FROM tipo_edital")
    cur.execute("DELETE FROM municipio")
    cur.execute("DELETE FROM estado")
    cur.execute("DELETE FROM pais")
    cur.execute("DELETE FROM subtema")
    cur.execute("DELETE FROM tema")

    def limpar(val):
        """Converte qualquer valor para string limpa ou None."""
        if val is None:
            return None
        import math
        try:
            if isinstance(val, float) and math.isnan(val):
                return None
        except Exception:
            pass
        s = str(val).strip()
        return None if s in ("", "-", "nan", "None", "NaN", "<NA>") else s

    def upsert(table, nome):
        nome = limpar(nome)
        if not nome:
            return None
        cur.execute(f"SELECT id FROM {table} WHERE nome = %s", (nome,))
        row = cur.fetchone()
        if row:
            return row[0]
        cur.execute(f"INSERT INTO {table} (nome) VALUES (%s) RETURNING id", (nome,))
        return cur.fetchone()[0]

    tema_map, subtema_map, pais_map, estado_map = {}, {}, {}, {}
    municipio_map, tipo_map, unidade_map, fonte_map, servico_map = {}, {}, {}, {}, {}

    for _, row in df.iterrows():
        tema_id = None
        v = limpar(row.get("tema"))
        if v:
            tema_id = tema_map.get(v) or upsert("tema", v)
            tema_map[v] = tema_id

        subtema_id = None
        v = limpar(row.get("subtema"))
        if v:
            key = (tema_id, v)
            if key not in subtema_map:
                cur.execute("SELECT id FROM subtema WHERE tema_id IS NOT DISTINCT FROM %s AND nome = %s", (tema_id, v))
                found = cur.fetchone()
                if found:
                    subtema_map[key] = found[0]
                else:
                    cur.execute("INSERT INTO subtema (tema_id, nome) VALUES (%s, %s) RETURNING id", (tema_id, v))
                    subtema_map[key] = cur.fetchone()[0]
            subtema_id = subtema_map[key]

        pais_id = None
        v = limpar(row.get("pais"))
        if v:
            pais_id = pais_map.get(v) or upsert("pais", v)
            pais_map[v] = pais_id

        estado_id = None
        v = limpar(row.get("estado"))
        if v:
            if v not in estado_map:
                cur.execute("SELECT id FROM estado WHERE nome = %s", (v,))
                found = cur.fetchone()
                if found:
                    estado_map[v] = found[0]
                else:
                    cur.execute("INSERT INTO estado (pais_id, nome) VALUES (%s, %s) RETURNING id", (pais_id, v))
                    estado_map[v] = cur.fetchone()[0]
            estado_id = estado_map[v]

        municipio_id = None
        v = limpar(row.get("municipio"))
        if v:
            key = (estado_id, v)
            if key not in municipio_map:
                cur.execute("SELECT id FROM municipio WHERE estado_id IS NOT DISTINCT FROM %s AND nome = %s", (estado_id, v))
                found = cur.fetchone()
                if found:
                    municipio_map[key] = found[0]
                else:
                    cur.execute("INSERT INTO municipio (estado_id, nome) VALUES (%s, %s) RETURNING id", (estado_id, v))
                    municipio_map[key] = cur.fetchone()[0]
            municipio_id = municipio_map[key]

        tipo_id = None
        v = limpar(row.get("tipo_edital"))
        if v:
            tipo_id = tipo_map.get(v) or upsert("tipo_edital", v)
            tipo_map[v] = tipo_id

        unidade_id = None
        v = limpar(row.get("unidade"))
        if v:
            unidade_id = unidade_map.get(v) or upsert("unidade", v)
            unidade_map[v] = unidade_id

        fonte_id = None
        v = limpar(row.get("fonte_dado"))
        if v:
            fonte_id = fonte_map.get(v) or upsert("fonte_dado", v)
            fonte_map[v] = fonte_id

        def safe_float(v):
            if v is None:
                return None
            try:
                if pd.isna(v):
                    return None
            except Exception:
                pass
            s = str(v).strip()
            if s in ("", "-", "nan", "None", "NaN"):
                return None
            # Trata formato brasileiro: 1.234.567,89 → 1234567.89
            if "," in s and "." in s:
                s = s.replace(".", "").replace(",", ".")
            elif "," in s:
                s = s.replace(",", ".")
            # Remove espaços internos (ex: "5 77.413,08")
            s = s.replace(" ", "")
            try:
                return float(s)
            except Exception:
                return None

        cur.execute("""
            INSERT INTO edital (
                tema_id, subtema_id, pais_id, estado_id, municipio_id, nome_edital,
                descricao, esforco, unidade_id, prazo_meses, tipo_edital_id,
                codigo_planilha, fonte_dado_id, observacao, custo_execucao,
                data_edital, metodo_calculo, valor_min, valor_max
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id
        """, (
            tema_id, subtema_id, pais_id, estado_id, municipio_id,
            limpar(row.get("nome_edital")), limpar(row.get("descricao")), limpar(row.get("esforco")),
            unidade_id, safe_float(row.get("prazo_meses")), tipo_id,
            limpar(row.get("codigo_planilha")), fonte_id, limpar(row.get("observacao")),
            safe_float(row.get("custo_execucao")), limpar(row.get("data_edital")),
            limpar(row.get("metodo_calculo")),
            safe_float(row.get("valor_min")), safe_float(row.get("valor_max")),
        ))
        edital_id = cur.fetchone()[0]

        for serv in normalizar_servicos(row.get("servicos")):
            servico_id = servico_map.get(serv) or upsert("servico", serv)
            servico_map[serv] = servico_id
            cur.execute("INSERT INTO edital_servico (edital_id, servico_id) VALUES (%s, %s) ON CONFLICT DO NOTHING", (edital_id, servico_id))

    conn.commit()
    conn.close()


def enviar_email(destinatarios, assunto: str, corpo_html: str):
    if not destinatarios:
        return False, "Nenhum destinatário informado."

    if not all([SMTP_SERVER, SMTP_PORT, SMTP_USER, SMTP_PASSWORD, EMAIL_FROM]):
        return False, "SMTP não configurado. Defina SMTP_SERVER, SMTP_PORT, SMTP_USER, SMTP_PASSWORD e EMAIL_FROM."

    try:
        msg = MIMEMultipart()
        msg["From"] = EMAIL_FROM
        msg["To"] = ", ".join(destinatarios)
        msg["Subject"] = assunto
        msg.attach(MIMEText(corpo_html, "html", "utf-8"))

        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(SMTP_USER, SMTP_PASSWORD)
            server.sendmail(EMAIL_FROM, destinatarios, msg.as_string())

        return True, None
    except Exception as e:
        return False, str(e)


def _base_email(conteudo_interno: str) -> str:
    return f"""
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head><meta charset="UTF-8"></head>
    <body style="margin:0;padding:0;background:#f4f6f9;font-family:Arial,sans-serif;">
      <table width="100%" cellpadding="0" cellspacing="0" style="background:#f4f6f9;padding:32px 0;">
        <tr><td align="center">
          <table width="600" cellpadding="0" cellspacing="0" style="background:#ffffff;border-radius:8px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,0.08);">
            <!-- Cabeçalho -->
            <tr>
              <td style="background:#1e3a8a;padding:24px 32px;">
                <p style="margin:0;color:#ffffff;font-size:18px;font-weight:bold;">FGV PMO</p>
                <p style="margin:4px 0 0;color:#93c5fd;font-size:13px;">Portal de Consulta de Editais</p>
              </td>
            </tr>
            <!-- Conteúdo -->
            <tr>
              <td style="padding:32px;">
                {conteudo_interno}
              </td>
            </tr>
            <!-- Rodapé -->
            <tr>
              <td style="background:#f8fafc;padding:16px 32px;border-top:1px solid #e2e8f0;">
                <p style="margin:0;color:#94a3b8;font-size:12px;">
                  Este é um e-mail automático enviado pelo Portal de Editais FGV PMO. Por favor, não responda diretamente a esta mensagem.
                </p>
              </td>
            </tr>
          </table>
        </td></tr>
      </table>
    </body>
    </html>
    """


def _linha_info(label: str, valor: str) -> str:
    return f"""
    <tr>
      <td style="padding:8px 12px;font-size:13px;color:#64748b;font-weight:600;width:160px;vertical-align:top;">{label}</td>
      <td style="padding:8px 12px;font-size:13px;color:#0f172a;vertical-align:top;">{valor}</td>
    </tr>"""


def _badge_status(status: str) -> str:
    cores = {
        "PENDENTE":    ("#fef3c7", "#92400e"),
        "EM ANÁLISE":  ("#dbeafe", "#1e40af"),
        "CONCLUÍDA":   ("#dcfce7", "#166534"),
        "RECUSADA":    ("#fee2e2", "#991b1b"),
    }
    bg, fg = cores.get(status, ("#f1f5f9", "#475569"))
    return f'<span style="background:{bg};color:{fg};padding:4px 12px;border-radius:20px;font-size:12px;font-weight:600;">{status}</span>'


def enviar_email_nova_solicitacao_para_admins(tema: str, descricao: str, solicitante: str, perfil: str, solicitacao_id: int):
    import html as _html
    emails_admin = buscar_emails_admins()
    if not emails_admin:
        return False, "Nenhum ADMIN com e-mail cadastrado."

    assunto = f"[FGV PMO] Nova solicitação de busca de edital — #{solicitacao_id}"
    conteudo = f"""
        <h2 style="margin:0 0 8px;color:#1e3a8a;font-size:20px;">Nova solicitação recebida</h2>
        <p style="margin:0 0 24px;color:#64748b;font-size:14px;">
          Uma nova solicitação de busca de edital foi registrada no Portal e aguarda análise.
        </p>
        <table width="100%" cellpadding="0" cellspacing="0"
               style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:6px;margin-bottom:24px;">
          {_linha_info("Nº da solicitação", f"#{solicitacao_id}")}
          {_linha_info("Data", agora_str())}
          {_linha_info("Solicitante", _html.escape(solicitante))}
          {_linha_info("Perfil", perfil)}
          {_linha_info("Tema solicitado", _html.escape(tema))}
          {_linha_info("Descrição", _html.escape(descricao) if descricao else "—")}
          {_linha_info("Status atual", _badge_status("PENDENTE"))}
        </table>
        <p style="margin:0;font-size:14px;color:#475569;">
          Acesse o Portal para analisar e atualizar o status desta solicitação.
        </p>
    """
    return enviar_email(emails_admin, assunto, _base_email(conteudo))


def enviar_email_atualizacao_status_para_admins(solicitacao_id: int, tema: str, solicitante: str, novo_status: str):
    import html as _html
    emails_admin = buscar_emails_admins()
    if not emails_admin:
        return False, "Nenhum ADMIN com e-mail cadastrado."

    assunto = f"[FGV PMO] Atualização de status — Solicitação #{solicitacao_id}"
    conteudo = f"""
        <h2 style="margin:0 0 8px;color:#1e3a8a;font-size:20px;">Status de solicitação atualizado</h2>
        <p style="margin:0 0 24px;color:#64748b;font-size:14px;">
          O status da solicitação abaixo foi atualizado no Portal.
        </p>
        <table width="100%" cellpadding="0" cellspacing="0"
               style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:6px;margin-bottom:24px;">
          {_linha_info("Nº da solicitação", f"#{solicitacao_id}")}
          {_linha_info("Data da atualização", agora_str())}
          {_linha_info("Solicitante", _html.escape(solicitante))}
          {_linha_info("Tema solicitado", _html.escape(tema))}
          {_linha_info("Novo status", _badge_status(novo_status))}
        </table>
        <p style="margin:0;font-size:14px;color:#475569;">
          Acesse o Portal para gerenciar as solicitações.
        </p>
    """
    return enviar_email(emails_admin, assunto, _base_email(conteudo))


def to_excel_bytes(df: pd.DataFrame) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Consulta")
    return output.getvalue()


def formatar_numero(valor):
    try:
        return f"{float(valor):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return str(valor)


def pode_solicitar(perfil):
    return perfil in ["ADMIN", "PMO", "COORDENADOR"]


def pode_ver_solicitacoes(perfil):
    return perfil in ["ADMIN", "PMO"]


def pode_gerenciar_usuarios(perfil):
    return perfil == "ADMIN"


def pode_substituir_base(perfil):
    return perfil in ["ADMIN", "PMO"]


def pode_baixar_arquivos(perfil):
    return perfil in ["ADMIN", "PMO", "COORDENADOR"]


# =========================================================
# SESSION
# =========================================================
def init_session():
    if "logado" not in st.session_state:
        st.session_state.logado = False
    if "usuario" not in st.session_state:
        st.session_state.usuario = None
    if "perfil" not in st.session_state:
        st.session_state.perfil = None
    if "email" not in st.session_state:
        st.session_state.email = None
    if "menu" not in st.session_state:
        st.session_state.menu = "Base de Prazos"
    if "tema_visual" not in st.session_state:
        st.session_state.tema_visual = "Light"


def logout():
    st.session_state.logado = False
    st.session_state.usuario = None
    st.session_state.perfil = None
    st.session_state.email = None
    st.session_state.menu = "Base de Prazos"
    st.rerun()


# =========================================================
# UI AUX
# =========================================================
def get_base64_logo():
    if not os.path.exists(LOGO_PATH):
        return ""
    with open(LOGO_PATH, "rb") as img_file:
        return base64.b64encode(img_file.read()).decode()
    
def get_base64_logo_completo():
    with open("assets/FGV_PMO_LOGO_COMPLETO.png", "rb") as img_file:
        return base64.b64encode(img_file.read()).decode()

def esconder_elementos_streamlit():
    """Remove elementos indesejados do Streamlit via JS."""
    st.markdown("""
    <script>
    (function() {
        function removeUnwanted() {
            // Remove pelo conteúdo de texto
            var allElements = document.querySelectorAll('button, span, div');
            allElements.forEach(function(el) {
                var txt = el.textContent || '';
                if (txt.includes('keyboard_double_arrow')) {
                    el.style.setProperty('display', 'none', 'important');
                    if (el.parentElement) {
                        el.parentElement.style.setProperty('display', 'none', 'important');
                    }
                }
            });
            // Remove botões no canto superior da página (fora da sidebar)
            var buttons = document.querySelectorAll('button');
            buttons.forEach(function(btn) {
                var rect = btn.getBoundingClientRect();
                // Botão no canto superior esquerdo absoluto (colapso da sidebar)
                if (rect.top < 60 && rect.left < 60 && rect.width < 60) {
                    btn.style.setProperty('display', 'none', 'important');
                }
            });
        }
        removeUnwanted();
        setTimeout(removeUnwanted, 500);
        setTimeout(removeUnwanted, 1500);
        var observer = new MutationObserver(function() { removeUnwanted(); });
        observer.observe(document.body, { childList: true, subtree: true });

        // Highlight active nav button
        function highlightActiveNav() {
            var menuName = window._activeMenu || '';
            var sidebar = document.querySelector('[data-testid="stSidebar"]');
            if (!sidebar) return;
            sidebar.querySelectorAll('.nav-menu button, div.nav-menu ~ div button').forEach(function(btn) {
                var txt = btn.innerText || '';
                // Remove previous active style
                btn.style.removeProperty('background');
                btn.style.removeProperty('color');
                btn.style.removeProperty('-webkit-text-fill-color');
                btn.style.removeProperty('border-left');
                btn.style.removeProperty('font-weight');
            });
        }
        setTimeout(highlightActiveNav, 300);
    })();
    </script>
    """, unsafe_allow_html=True)


def header_principal():
    esconder_elementos_streamlit()
    logo_b64 = get_base64_logo_completo()

    st.markdown(
        f"""
        <div class="header-full-width">
            <div class="header-inner">
                <div class="header-text-block">
                    <div class="header-title">Consulta e busca de novos editais</div>
                    <div class="header-subtitle">Para calculo de prazo completo, entrar em contato com FGV PMO</div>
                    <div class="header-profile"><b>Perfil ativo:</b> {st.session_state.usuario}</div>
                </div>
                <div class="header-logo-block">
                    <img src="data:image/png;base64,{logo_b64}" class="header-logo-full"/>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True
    )


def metric_card(titulo, valor, subtitulo=""):
    st.markdown(f"""
    <div class="metric-card">
        <div class="metric-title">{titulo}</div>
        <div class="metric-value">{valor}</div>
        <div class="metric-sub">{subtitulo}</div>
    </div>
    """, unsafe_allow_html=True)


# =========================================================
# LOGIN
# =========================================================
def tela_login():
    st.markdown("""
    <style>
    .login-page-wrap {
        min-height: 0vh;
        display: flex;
        justify-content: center;
        align-items: center;
        padding: 0px;
    }

    .login-panel {
        width: 100%;
        max-width: 420px;
    }

    .login-brand {
        text-align: center;
        margin-bottom: 18px;
    }

    .login-brand img {
        width: 220px;
        max-width: 100%;
        object-fit: contain;
        margin-bottom: 10px;
    }

    .login-card {
        padding: 5px;
    }

    .login-title {
        text-align: center;
        font-size: 1.9rem;
        font-weight: 800;
        margin-bottom: 8px;
    }

    .login-subtitle {
        text-align: center;
        font-size: 0.98rem;
        margin-bottom: 22px;
        opacity: 0.9;
    }

    .login-footer {
        margin-top: 16px;
        text-align: center;
        font-size: 0.88rem;
        opacity: 0.85;
    }
    </style>
    """, unsafe_allow_html=True)

    st.markdown('<div class="login-page-wrap">', unsafe_allow_html=True)
    st.markdown('<div class="login-panel">', unsafe_allow_html=True)

    st.markdown("""
    <div class="login-brand">
        <img src="data:image/png;base64,{}" />
    </div>
    """.format(get_base64_logo()), unsafe_allow_html=True)

    st.markdown("""
    <div class="login-card">
        <div class="login-title">Portal de Consulta de Editais</div>
        <div class="login-subtitle">FGV PMO</div>
    """, unsafe_allow_html=True)

    with st.form("form_login", clear_on_submit=False):
        usuario = st.text_input("Usuário")
        senha = st.text_input("Senha", type="password")
        entrar = st.form_submit_button("Entrar", use_container_width=True)

    st.markdown("""
        <div class="login-footer">
            Acesso restrito a usuários autorizados
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("</div>", unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)

    if entrar:
        try:
            user = autenticar(usuario, senha)
        except PermissionError as pe:
            st.error(str(pe))
            st.stop()
        if user:
            st.session_state.logado = True
            st.session_state.usuario = user["username"]
            st.session_state.perfil = user["perfil"]
            st.rerun()
        else:
            st.error("Usuário ou senha inválidos.")


# =========================================================
# SIDEBAR
# =========================================================
def menu_sidebar():
    with st.sidebar:
        tema_atual = st.session_state.get("tema_visual", "Light")

        st.markdown('<div class="theme-toggle-mini">', unsafe_allow_html=True)

        col1, col2, col3 = st.columns([1, 1, 2])

        with col1:
            if st.button("☀", key="btn_light", use_container_width=True):
                if tema_atual != "Light":
                    st.session_state.tema_visual = "Light"
                    st.rerun()

        with col2:
            if st.button("☾", key="btn_dark", use_container_width=True):
                if tema_atual != "Dark":
                    st.session_state.tema_visual = "Dark"
                    st.rerun()

        st.markdown("</div>", unsafe_allow_html=True)

        st.markdown("""
        <div class="sidebar-logo-wrap">
            <img src="data:image/png;base64,{}" class="sidebar-logo-img"/>
        </div>
        """.format(get_base64_logo()), unsafe_allow_html=True)

        st.markdown("## Portal de Editais")
        st.markdown(f"**Usuário:** {st.session_state.usuario}")

        perfil = st.session_state.perfil
        opcoes = ["Base de Prazos"]

        if perfil in ("ADMIN", "PMO"):
            opcoes.append("Análise de Prazos")

        opcoes.append("Projetos Concluídos")

        if perfil in ("ADMIN", "PMO", "COORDENADOR"):
            # Badge de pendentes para ADMIN/PMO
            if perfil in ("ADMIN", "PMO"):
                try:
                    df_pend = listar_solicitacoes()
                    n_pend = len(df_pend[df_pend["status"] == "PENDENTE"])
                    opcoes.append(f"Solicitações ({n_pend})" if n_pend > 0 else "Solicitações")
                except Exception:
                    opcoes.append("Solicitações")
            else:
                opcoes.append("Solicitações")

        if perfil in ("ADMIN", "PMO"):
            opcoes.append("Base de dados")

        opcoes.append("Minha conta")

        if perfil == "ADMIN":
            opcoes.append("Usuários")


        menu_atual = st.session_state.menu
        # Injeta JS para estilizar botão ativo
        st.markdown(f"""
        <script>
        (function applyActiveNav() {{
            var activeText = {repr(menu_atual)};
            var sidebar = document.querySelector('[data-testid="stSidebar"]');
            if (!sidebar) {{ setTimeout(applyActiveNav, 100); return; }}
            sidebar.querySelectorAll('button').forEach(function(btn) {{
                var txt = (btn.innerText || '').trim();
                if (txt === activeText) {{
                    btn.style.setProperty('background', 'rgba(255,255,255,0.18)', 'important');
                    btn.style.setProperty('color', '#ffffff', 'important');
                    btn.style.setProperty('-webkit-text-fill-color', '#ffffff', 'important');
                    btn.style.setProperty('font-weight', '700', 'important');
                    btn.style.setProperty('border', '1px solid rgba(255,255,255,0.3)', 'important');
                }} else {{
                    btn.style.removeProperty('border');
                }}
            }});
        }})();
        setTimeout(function() {{
            var activeText = {repr(menu_atual)};
            var sidebar = document.querySelector('[data-testid="stSidebar"]');
            if (!sidebar) return;
            sidebar.querySelectorAll('button').forEach(function(btn) {{
                var txt = (btn.innerText || '').trim();
                if (txt === activeText) {{
                    btn.style.setProperty('background', 'rgba(255,255,255,0.18)', 'important');
                    btn.style.setProperty('color', '#ffffff', 'important');
                    btn.style.setProperty('-webkit-text-fill-color', '#ffffff', 'important');
                    btn.style.setProperty('font-weight', '700', 'important');
                    btn.style.setProperty('border', '1px solid rgba(255,255,255,0.3)', 'important');
                }}
            }});
        }}, 400);
        </script>
        """, unsafe_allow_html=True)
        for opcao in opcoes:
            if st.button(opcao, key=f"nav_{opcao}", use_container_width=True):
                if st.session_state.menu != opcao:
                    st.session_state.menu = opcao
                    st.rerun()

        st.markdown("---")
        if st.button("Sair", use_container_width=True, key="btn_sair_sidebar"):
            logout()


# =========================================================
# CONSULTA
# =========================================================
def pagina_consulta():
    header_principal()
    df = carregar_view()

    if df.empty:
        st.warning("A view 'vw_consulta_editais' não foi encontrada ou não possui dados.")
        return

    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].fillna("").astype(str)

    df.columns = [c.strip() for c in df.columns]

    def achar_coluna(preferidas):
        for p in preferidas:
            for c in df.columns:
                if c.lower() == p.lower():
                    return c
        return None

    col_tema = achar_coluna(["tema"])
    col_subtema = achar_coluna(["subtema"])
    col_estado = achar_coluna(["estado"])
    col_municipio = achar_coluna(["municipio", "município"])
    col_nome = achar_coluna(["nome", "denominacao", "denominação"])
    col_desc = achar_coluna(["descricao", "descrição"])
    col_codigo = achar_coluna(["codigo", "código", "codigo_planilha"])
    col_obs = achar_coluna(["observacao", "observação", "obs"])
    col_custo = achar_coluna(["custo", "valor", "custo_execucao"])
    col_prazo = achar_coluna(["prazo_meses", "prazo"])
    col_data = achar_coluna(["data_edital", "data edital", "data"])

    if col_custo:
        df[col_custo] = pd.to_numeric(df[col_custo], errors="coerce")
    if col_prazo:
        df[col_prazo] = pd.to_numeric(df[col_prazo], errors="coerce")
    if col_data:
        df[col_data] = pd.to_datetime(df[col_data], errors="coerce")


    st.subheader("Filtros de consulta")

    def opcoes(df_base, col):
        if not col:
            return ["Todos"]
        return ["Todos"] + sorted(df_base[col].dropna().replace("", pd.NA).dropna().unique().tolist())

    # Filtros em cascata: cada filtro restringe as opções dos seguintes
    filtrado = df.copy()

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        tema = st.selectbox("Tema", opcoes(filtrado, col_tema))
    if col_tema and tema != "Todos":
        filtrado = filtrado[filtrado[col_tema] == tema]

    with c2:
        subtema = st.selectbox("Subtema", opcoes(filtrado, col_subtema))
    if col_subtema and subtema != "Todos":
        filtrado = filtrado[filtrado[col_subtema] == subtema]

    with c3:
        estado = st.selectbox("Estado", opcoes(filtrado, col_estado))
    if col_estado and estado != "Todos":
        filtrado = filtrado[filtrado[col_estado] == estado]

    with c4:
        municipio = st.selectbox("Município", opcoes(filtrado, col_municipio))
    if col_municipio and municipio != "Todos":
        filtrado = filtrado[filtrado[col_municipio] == municipio]

    busca = st.text_input("Busca textual", placeholder="Nome, descrição, código...")

    c9, c10, c11, c12 = st.columns(4)
    custo_min = c9.number_input("Custo mínimo", min_value=0.0, value=0.0, step=1000.0, format="%.2f")
    custo_max = c10.number_input("Custo máximo", min_value=0.0, value=0.0, step=1000.0, format="%.2f")
    prazo_min = c11.number_input("Prazo mínimo (meses)", min_value=0.0, value=0.0, step=1.0, format="%.2f")
    prazo_max = c12.number_input("Prazo máximo (meses)", min_value=0.0, value=0.0, step=1.0, format="%.2f")
    st.markdown('</div>', unsafe_allow_html=True)

    if busca:
        texto_cols = [c for c in [col_nome, col_desc, col_codigo, col_obs] if c]
        if texto_cols:
            mask = False
            for c in texto_cols:
                mask = mask | filtrado[c].astype(str).str.contains(busca, case=False, na=False)
            filtrado = filtrado[mask]

    if col_custo:
        if custo_min > 0:
            filtrado = filtrado[filtrado[col_custo] >= custo_min]
        if custo_max > 0:
            filtrado = filtrado[filtrado[col_custo] <= custo_max]
    if col_prazo:
        if prazo_min > 0:
            filtrado = filtrado[filtrado[col_prazo] >= prazo_min]
        if prazo_max > 0:
            filtrado = filtrado[filtrado[col_prazo] <= prazo_max]

    total_registros = len(filtrado)
    total_temas = filtrado[col_tema].nunique() if col_tema else 0
    total_estados = filtrado[col_estado].nunique() if col_estado else 0
    custo_medio = filtrado[col_custo].mean() if col_custo and not filtrado.empty else 0

    m1, m2, m3, m4 = st.columns(4)
    with m1:
        metric_card("Registros filtrados", total_registros)
    with m2:
        metric_card("Temas", total_temas)
    with m3:
        metric_card("Estados", total_estados)
    with m4:
        metric_card("Custo médio", formatar_numero(custo_medio if pd.notna(custo_medio) else 0))


    st.subheader("Resultados")

    colunas_remover = [
        "id", "tipo_edital", "codigo_planilha", "fonte_dado", "metodo_calculo", "valor_min", "valor_max", "observacao"
    ]
    colunas_remover_existentes = [c for c in colunas_remover if c in filtrado.columns]
    df_exibicao = filtrado.drop(columns=colunas_remover_existentes)

    mapa_colunas = {
        "codigo": "Código",
        "nome": "Nome",
        "descricao": "Objetivo do Projeto",
        "tema": "Tema",
        "subtema": "Subtema",
        "pais": "País",
        "estado": "Estado",
        "municipio": "Município",
        "nome_edital": "Edital",
        "esforco": "Parâmetro utilizado para verificação do prazo",
        "unidade": "Unidade de Medida do Parâmetro ",
        "servicos": "Serviços",
        "custo_execucao": "Custo (R$)",
        "custo": "Custo (R$)",
        "prazo_meses": "Prazo de execução (meses)",
        "data_edital": "Data do edital"
    }
    df_exibicao = df_exibicao.rename(columns={k: v for k, v in mapa_colunas.items() if k in df_exibicao.columns})

    if "Data do edital" in df_exibicao.columns:
        df_exibicao["Data do edital"] = pd.to_datetime(df_exibicao["Data do edital"], errors="coerce").dt.strftime("%d/%m/%Y")
    if "Custo (R$)" in df_exibicao.columns:
        df_exibicao["Custo (R$)"] = df_exibicao["Custo (R$)"].apply(
            lambda x: f"R$ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".") if pd.notnull(x) else ""
        )

    # ── Paginação ──
    PAGE_SIZE = 50
    total = len(df_exibicao)
    n_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    if "pagina_consulta" not in st.session_state:
        st.session_state["pagina_consulta"] = 1
    # Reset se filtros mudaram e página ficou fora do range
    if st.session_state["pagina_consulta"] > n_pages:
        st.session_state["pagina_consulta"] = 1

    pg_atual = st.session_state["pagina_consulta"]
    inicio = (pg_atual - 1) * PAGE_SIZE
    fim = min(inicio + PAGE_SIZE, total)

    st.dataframe(df_exibicao.iloc[inicio:fim], use_container_width=True, hide_index=True)

    # Controles de paginação
    pg1, pg2, pg3, pg4, pg5 = st.columns([1, 1, 3, 1, 1])
    with pg1:
        if st.button("⟪ Primeira", use_container_width=True, disabled=pg_atual == 1):
            st.session_state["pagina_consulta"] = 1
            st.rerun()
    with pg2:
        if st.button("‹ Anterior", use_container_width=True, disabled=pg_atual == 1):
            st.session_state["pagina_consulta"] -= 1
            st.rerun()
    with pg3:
        st.markdown(
            f"<div style='text-align:center;padding:8px 0;font-size:13px;color:var(--ink-secondary);'>"
            f"Página <b>{pg_atual}</b> de <b>{n_pages}</b> &nbsp;·&nbsp; "
            f"Exibindo registros <b>{inicio+1}</b>–<b>{fim}</b> de <b>{total}</b></div>",
            unsafe_allow_html=True
        )
    with pg4:
        if st.button("Próxima ›", use_container_width=True, disabled=pg_atual == n_pages):
            st.session_state["pagina_consulta"] += 1
            st.rerun()
    with pg5:
        if st.button("Última ⟫", use_container_width=True, disabled=pg_atual == n_pages):
            st.session_state["pagina_consulta"] = n_pages
            st.rerun()

    if pode_baixar_arquivos(st.session_state.perfil):
        col_dl1, col_dl2 = st.columns([1, 5])
        with col_dl1:
            st.download_button(
                "Baixar CSV",
                data=filtrado.to_csv(index=False).encode("utf-8-sig"),
                file_name="consulta_editais.csv",
                mime="text/csv",
                use_container_width=True
            )
        with col_dl2:
            st.download_button(
                "Baixar Excel",
                data=to_excel_bytes(filtrado),
                file_name="consulta_editais.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=False
            )

    # ── Gráfico de evolução por tema ──
    if not filtrado.empty and "data_edital" in filtrado.columns and "tema" in filtrado.columns:
        try:
            import plotly.graph_objects as go
            df_graf = filtrado[["data_edital", "tema"]].copy()
            df_graf["data_edital"] = pd.to_datetime(df_graf["data_edital"], errors="coerce")
            df_graf = df_graf.dropna(subset=["data_edital"])
            df_graf["ano"] = df_graf["data_edital"].dt.year.astype(int)
            df_graf = df_graf[df_graf["ano"] >= 2015]
            if not df_graf.empty:
                pivot = df_graf.groupby(["ano", "tema"]).size().reset_index(name="n")
                temas_graf = pivot.groupby("tema")["n"].sum().nlargest(8).index.tolist()
                pivot = pivot[pivot["tema"].isin(temas_graf)]
                cores = ["#1d6fc4","#10b981","#f59e0b","#ef4444","#8b5cf6",
                         "#06b6d4","#ec4899","#84cc16"]
                fig = go.Figure()
                for i, tema_g in enumerate(temas_graf):
                    d = pivot[pivot["tema"] == tema_g].sort_values("ano")
                    fig.add_trace(go.Scatter(
                        x=d["ano"].tolist(), y=d["n"].tolist(),
                        name=tema_g, mode="lines+markers",
                        line=dict(width=2, color=cores[i % len(cores)]),
                        marker=dict(size=6),
                    ))
                fig.update_layout(
                    title="Evolução de editais por tema ao longo do tempo",
                    xaxis_title="Ano", yaxis_title="Nº de editais",
                    height=380, template="plotly_white",
                    legend=dict(orientation="h", yanchor="bottom", y=-0.4),
                    margin=dict(t=50, b=100, l=40, r=20),
                    xaxis=dict(tickmode="linear", dtick=1),
                )
                st.plotly_chart(fig, use_container_width=True)
        except ImportError:
            pass




# =========================================================
# SOLICITAÇÕES
# =========================================================
def pagina_solicitacoes():
    header_principal()

    if pode_solicitar(st.session_state.perfil):
        st.markdown('<div class="section-card">', unsafe_allow_html=True)
        st.subheader("Solicitar busca de novos editais")

        with st.form("form_solicitacao_tema", clear_on_submit=True):
            tema = st.text_input("Tema da pesquisa")
            descricao = st.text_area("Descrição complementar", placeholder="Explique melhor o tema, palavras-chave, região, observações...")
            enviar = st.form_submit_button("Enviar solicitação")

        if enviar:
            if not tema.strip():
                st.warning("Informe o tema da pesquisa.")
            else:
                solicitacao_id = inserir_solicitacao(
                    tema=tema,
                    descricao=descricao,
                    solicitante=st.session_state.usuario,
                    perfil=st.session_state.perfil
                )
                ok_email, msg_email = enviar_email_nova_solicitacao_para_admins(
                    tema=tema,
                    descricao=descricao,
                    solicitante=st.session_state.usuario,
                    perfil=st.session_state.perfil,
                    solicitacao_id=solicitacao_id
                )
                if ok_email:
                    st.success("Solicitação registrada com sucesso e notificação enviada aos administradores.")
                else:
                    st.success("Solicitação registrada com sucesso.")
                    st.info(f"Aviso sobre e-mail: {msg_email}")
                st.rerun()

        st.markdown('</div>', unsafe_allow_html=True)

    if pode_ver_solicitacoes(st.session_state.perfil):
        st.markdown('<div class="section-card">', unsafe_allow_html=True)
        st.subheader("Solicitações recebidas")

        df_sol = listar_solicitacoes()
        if df_sol.empty:
            st.info("Nenhuma solicitação cadastrada.")
        else:
            st.dataframe(df_sol, use_container_width=True, hide_index=True)
            st.markdown("### Atualizar status")
            c1, c2 = st.columns(2)
            with c1:
                ids = df_sol["id"].tolist()
                solicitacao_id = st.selectbox("ID da solicitação", ids)
            with c2:
                novo_status = st.selectbox("Novo status", ["PENDENTE", "EM ANÁLISE", "CONCLUÍDA", "RECUSADA"])

            if st.button("Salvar status"):
                dados_sol = obter_solicitacao_por_id(solicitacao_id)
                if not dados_sol:
                    st.error("Solicitação não encontrada.")
                else:
                    _, tema_solicitado, _, solicitante, _, _, _ = dados_sol
                    atualizar_status_solicitacao(solicitacao_id, novo_status)
                    ok_email, msg_email = enviar_email_atualizacao_status_para_admins(
                        solicitacao_id=solicitacao_id,
                        tema=tema_solicitado,
                        solicitante=solicitante,
                        novo_status=novo_status
                    )
                    if ok_email:
                        st.success("Status atualizado e e-mail enviado aos administradores.")
                    else:
                        st.success("Status atualizado.")
                        st.info(f"Aviso sobre e-mail: {msg_email}")
                    st.rerun()

        st.markdown('</div>', unsafe_allow_html=True)
    # ── Histórico próprio (todos os perfis veem suas solicitações) ──
    st.markdown('<div class="section-card">', unsafe_allow_html=True)
    st.subheader("Minhas solicitações")
    df_todas = listar_solicitacoes()
    df_minhas = df_todas[df_todas["solicitante"].str.upper() == st.session_state.usuario.upper()].copy()
    if df_minhas.empty:
        st.info("Você ainda não possui solicitações registradas.")
    else:
        # Badges de status coloridos
        def badge_status(s):
            cores = {"PENDENTE": ("#fef3c7","#92400e"), "EM ANÁLISE": ("#dbeafe","#1e40af"),
                     "CONCLUÍDA": ("#dcfce7","#166534"), "RECUSADA": ("#fee2e2","#991b1b")}
            bg, fg = cores.get(s, ("#f1f5f9","#475569"))
            return f'<span style="background:{bg};color:{fg};padding:2px 10px;border-radius:999px;font-size:11px;font-weight:600;">{s}</span>'

        st.markdown(f"**{len(df_minhas)}** solicitação(ões) encontrada(s)")
        for _, row in df_minhas.iterrows():
            with st.expander(f"#{row['id']} — {row['tema_solicitado']}  |  {row['data_solicitacao'][:10]}"):
                col_a, col_b = st.columns([2,1])
                with col_a:
                    st.markdown(f"**Tema:** {row['tema_solicitado']}")
                    if row.get('descricao'):
                        st.markdown(f"**Descrição:** {row['descricao']}")
                    st.markdown(f"**Data:** {row['data_solicitacao']}")
                with col_b:
                    st.markdown(f"**Status:**", unsafe_allow_html=False)
                    st.markdown(badge_status(row['status']), unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)


# =========================================================
# BASE DE DADOS
# =========================================================
def _inserir_edital_individual(tema, subtema, pais, estado, municipio,
                                nome_edital, descricao, servicos, esforco,
                                unidade, prazo_meses, custo_execucao,
                                valor_min, valor_max, data_edital,
                                codigo_planilha, observacao):
    """Insere um único edital na base mantendo a normalização."""
    conn = get_conn()
    cur = conn.cursor()

    def upsert(table, nome):
        if not nome:
            return None
        cur.execute(f"SELECT id FROM {table} WHERE nome = %s", (nome,))
        row = cur.fetchone()
        if row:
            return row[0]
        cur.execute(f"INSERT INTO {table} (nome) VALUES (%s) RETURNING id", (nome,))
        return cur.fetchone()[0]

    # Dimensões
    tema_id = upsert("tema", tema)

    subtema_id = None
    if subtema:
        cur.execute("SELECT id FROM subtema WHERE tema_id = %s AND nome = %s", (tema_id, subtema))
        row = cur.fetchone()
        if row:
            subtema_id = row[0]
        else:
            cur.execute("INSERT INTO subtema (tema_id, nome) VALUES (%s, %s) RETURNING id", (tema_id, subtema))
            subtema_id = cur.fetchone()[0]

    pais_id = upsert("pais", pais)

    estado_id = None
    if estado:
        cur.execute("SELECT id FROM estado WHERE nome = %s", (estado,))
        row = cur.fetchone()
        if row:
            estado_id = row[0]
        else:
            cur.execute("INSERT INTO estado (pais_id, nome) VALUES (%s, %s) RETURNING id", (pais_id, estado))
            estado_id = cur.fetchone()[0]

    municipio_id = None
    if municipio:
        cur.execute("SELECT id FROM municipio WHERE estado_id IS NOT DISTINCT FROM %s AND nome = %s", (estado_id, municipio))
        row = cur.fetchone()
        if row:
            municipio_id = row[0]
        else:
            cur.execute("INSERT INTO municipio (estado_id, nome) VALUES (%s, %s) RETURNING id", (estado_id, municipio))
            municipio_id = cur.fetchone()[0]

    unidade_id = upsert("unidade", unidade)

    def safe_float(v):
        try:
            return float(v) if v else None
        except Exception:
            return None

    cur.execute("""
        INSERT INTO edital (
            tema_id, subtema_id, pais_id, estado_id, municipio_id,
            nome_edital, descricao, esforco, unidade_id, prazo_meses,
            codigo_planilha, observacao, custo_execucao,
            data_edital, valor_min, valor_max
        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        RETURNING id
    """, (
        tema_id, subtema_id, pais_id, estado_id, municipio_id,
        nome_edital, descricao, esforco, unidade_id,
        safe_float(prazo_meses), codigo_planilha, observacao,
        safe_float(custo_execucao), data_edital,
        safe_float(valor_min), safe_float(valor_max),
    ))
    edital_id = cur.fetchone()[0]

    # Serviços
    if servicos:
        for serv in normalizar_servicos(servicos):
            servico_id = None
            cur.execute("SELECT id FROM servico WHERE nome = %s", (serv,))
            row = cur.fetchone()
            if row:
                servico_id = row[0]
            else:
                cur.execute("INSERT INTO servico (nome) VALUES (%s) RETURNING id", (serv,))
                servico_id = cur.fetchone()[0]
            cur.execute("INSERT INTO edital_servico (edital_id, servico_id) VALUES (%s, %s) ON CONFLICT DO NOTHING",
                       (edital_id, servico_id))

    conn.commit()
    conn.close()
    return edital_id


def pagina_base():
    header_principal()
    df = carregar_view()

    st.markdown('<div class="section-card">', unsafe_allow_html=True)
    st.subheader("Visualização da base")
    if df.empty:
        st.warning("A view 'vw_consulta_editais' não foi encontrada ou não possui dados.")
    else:
        st.dataframe(df.head(500), use_container_width=True, hide_index=True)
        st.caption("Exibindo até 500 linhas para visualização.")
    st.markdown('</div>', unsafe_allow_html=True)

    if pode_substituir_base(st.session_state.perfil):
        st.markdown('<div class="section-card">', unsafe_allow_html=True)
        st.subheader("Substituir base de dados")
        st.info("Área reservada para ADMIN e PMO. O upload substitui todos os editais da base pelos da nova planilha. Usuários e solicitações não são afetados.")
        arquivo = st.file_uploader("Selecione uma planilha", type=["xlsx", "xls", "csv"])
        if arquivo is not None:
            st.success(f"Arquivo carregado: {arquivo.name}")
            if st.button("Processar e atualizar base", type="primary"):
                with st.spinner("Processando planilha e atualizando o banco..."):
                    try:
                        processar_upload_planilha(arquivo)
                        st.success("Base atualizada com sucesso! Recarregue a página para ver os novos dados.")
                        st.cache_data.clear()
                    except Exception as e:
                        logger.error("Erro ao processar planilha: %s", e)
                        st.error("Erro ao processar a planilha. Verifique o formato do arquivo e tente novamente.")
        st.markdown('</div>', unsafe_allow_html=True)

    # ── Inclusão de edital individual ──
    if pode_substituir_base(st.session_state.perfil):
        st.markdown('<div class="section-card">', unsafe_allow_html=True)
        st.subheader("Incluir novo edital individualmente")
        st.info("Adicione um edital diretamente na base sem precisar substituir a planilha inteira.")

        conn_ref = get_conn()
        try:
            temas_ref = pd.read_sql_query("SELECT DISTINCT nome FROM tema ORDER BY nome", conn_ref)["nome"].tolist()
            estados_ref = pd.read_sql_query("SELECT DISTINCT nome FROM estado ORDER BY nome", conn_ref)["nome"].tolist()
            unidades_ref = pd.read_sql_query("SELECT DISTINCT nome FROM unidade ORDER BY nome", conn_ref)["nome"].tolist()
        except Exception:
            temas_ref, estados_ref, unidades_ref = [], [], []
        finally:
            conn_ref.close()

        with st.form("form_novo_edital", clear_on_submit=True):
            st.markdown("**Identificação**")
            c1, c2 = st.columns(2)
            with c1:
                ne_tema = st.selectbox("Tema *", [""] + temas_ref)
                ne_subtema = st.text_input("Subtema *")
                ne_nome_edital = st.text_input("Nome do edital *")
            with c2:
                ne_pais = st.text_input("País", value="Brasil")
                ne_estado = st.selectbox("Estado", [""] + estados_ref)
                ne_municipio = st.text_input("Município")

            st.markdown("**Descrição**")
            ne_descricao = st.text_area("Descrição / Objetivo do edital", height=100)
            ne_servicos = st.text_input("Serviços (separados por vírgula)")

            st.markdown("**Dados técnicos**")
            c3, c4, c5 = st.columns(3)
            with c3:
                ne_esforco = st.number_input("Esforço", min_value=0.0, step=0.1, format="%.2f")
                ne_unidade = st.selectbox("Unidade", [""] + unidades_ref)
            with c4:
                ne_prazo = st.number_input("Prazo (meses)", min_value=0.0, step=0.5, format="%.1f")
                ne_custo = st.number_input("Custo de execução (R$)", min_value=0.0, step=1000.0, format="%.2f")
            with c5:
                ne_valor_min = st.number_input("Valor mínimo (R$)", min_value=0.0, step=1000.0, format="%.2f")
                ne_valor_max = st.number_input("Valor máximo (R$)", min_value=0.0, step=1000.0, format="%.2f")

            st.markdown("**Outros**")
            c6, c7 = st.columns(2)
            with c6:
                ne_data = st.text_input("Data do edital (AAAA-MM)", placeholder="Ex: 2024-03")
                ne_codigo = st.text_input("Código planilha")
            with c7:
                ne_obs = st.text_area("Observações", height=80)

            salvar_edital = st.form_submit_button("Salvar edital", type="primary")

        if salvar_edital:
            if not ne_tema or not ne_subtema.strip() or not ne_nome_edital.strip():
                st.warning("Preencha ao menos Tema, Subtema e Nome do edital.")
            else:
                try:
                    _inserir_edital_individual(
                        tema=ne_tema, subtema=ne_subtema.strip(),
                        pais=ne_pais.strip() or "Brasil", estado=ne_estado or None,
                        municipio=ne_municipio.strip() or None,
                        nome_edital=ne_nome_edital.strip(),
                        descricao=ne_descricao.strip() or None,
                        servicos=ne_servicos.strip() or None,
                        esforco=str(ne_esforco) if ne_esforco > 0 else None,
                        unidade=ne_unidade or None,
                        prazo_meses=ne_prazo if ne_prazo > 0 else None,
                        custo_execucao=ne_custo if ne_custo > 0 else None,
                        valor_min=ne_valor_min if ne_valor_min > 0 else None,
                        valor_max=ne_valor_max if ne_valor_max > 0 else None,
                        data_edital=ne_data.strip() or None,
                        codigo_planilha=ne_codigo.strip() or None,
                        observacao=ne_obs.strip() or None,
                    )
                    st.success(f"Edital '{ne_nome_edital}' incluído com sucesso!")
                    st.cache_data.clear()
                except Exception as e:
                    logger.error("Erro ao incluir edital: %s", e)
                    st.error("Erro ao salvar o edital. Tente novamente ou contate o administrador.")

        st.markdown('</div>', unsafe_allow_html=True)


# =========================================================
# MINHA CONTA
# =========================================================
def pagina_minha_conta():
    header_principal()

    st.markdown('<div class="section-card">', unsafe_allow_html=True)
    st.subheader("Minha conta")
    st.write(f"**Usuário:** {st.session_state.usuario}")
    st.write(f"**Perfil:** {st.session_state.perfil}")
    st.write(f"**E-mail:** {st.session_state.email or '-'}")
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="section-card">', unsafe_allow_html=True)
    st.subheader("Alterar senha")
    with st.form("form_alterar_senha", clear_on_submit=True):
        senha_atual = st.text_input("Senha atual", type="password")
        nova_senha = st.text_input("Nova senha", type="password")
        confirmar_senha = st.text_input("Confirmar nova senha", type="password")
        salvar = st.form_submit_button("Salvar nova senha")

    if salvar:
        if not senha_atual or not nova_senha or not confirmar_senha:
            st.warning("Preencha todos os campos.")
        elif nova_senha != confirmar_senha:
            st.error("A confirmação da nova senha não confere.")
        elif len(nova_senha) < 6:
            st.error("A nova senha deve ter pelo menos 6 caracteres.")
        else:
            ok, msg = alterar_senha_usuario(st.session_state.usuario, senha_atual, nova_senha)
            if ok:
                st.success(msg)
            else:
                st.error(msg)

    st.markdown('</div>', unsafe_allow_html=True)


# =========================================================
# USUÁRIOS
# =========================================================
def pagina_usuarios():
    header_principal()

    if not pode_gerenciar_usuarios(st.session_state.perfil):
        st.error("Acesso restrito ao perfil ADMIN.")
        return

    st.markdown('<div class="section-card">', unsafe_allow_html=True)
    st.subheader("Criar novo usuário")

    with st.form("form_novo_usuario", clear_on_submit=True):
        novo_user = st.text_input("Usuário")
        novo_email = st.text_input("E-mail")
        nova_senha = st.text_input("Senha", type="password")
        perfil = st.selectbox("Perfil", ["ADMIN", "PMO", "COORDENADOR", "GERAL"])
        criar = st.form_submit_button("Criar usuário")

    if criar:
        if not novo_user.strip() or not nova_senha.strip():
            st.warning("Preencha usuário e senha.")
        elif not novo_email.strip():
            st.warning("Preencha o e-mail do usuário.")
        else:
            ok_senha, msg_senha = validar_senha(nova_senha)
            if not ok_senha:
                st.warning(msg_senha)
            else:
                try:
                    criar_usuario(novo_user, novo_email, nova_senha, perfil)
                    st.success("Usuário criado com sucesso.")
                    st.rerun()
                except psycopg2.errors.UniqueViolation:
                    st.error("Já existe um usuário com esse nome ou e-mail.")
                except Exception as e:
                    logger.error("Erro ao criar usuario: %s", e)
                    st.error("Erro ao criar o usuário. Verifique se o nome ou e-mail já existe.")

    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="section-card">', unsafe_allow_html=True)
    st.subheader("Usuários cadastrados")

    df_users = listar_usuarios()
    if df_users.empty:
        st.info("Nenhum usuário cadastrado.")
    else:
        df_exib = df_users.copy()
        df_exib["ativo"] = df_exib["ativo"].map({1: "Sim", 0: "Não"})
        st.dataframe(df_exib, use_container_width=True, hide_index=True)

        st.markdown("### Gerenciar usuário")
        c1, c2 = st.columns(2)
        with c1:
            user_id = st.selectbox(
                "Usuário",
                df_users["id"].tolist(),
                format_func=lambda x: f"{x} - {df_users.loc[df_users['id'] == x, 'username'].values[0]}"
            )
        with c2:
            status = st.selectbox("Novo status", ["Ativo", "Inativo"])

        c3, c4 = st.columns(2)
        with c3:
            if st.button("Salvar alteração de status", use_container_width=True):
                alterar_status_usuario(user_id, 1 if status == "Ativo" else 0)
                st.success("Status atualizado com sucesso.")
                st.rerun()

        with c4:
            if st.button("Excluir usuário", use_container_width=True):
                username_selecionado = df_users.loc[df_users["id"] == user_id, "username"].values[0]
                if username_selecionado.upper() == st.session_state.usuario.upper():
                    st.error("Você não pode excluir o próprio usuário logado.")
                elif username_selecionado.upper() == "ADMIN":
                    st.error("Não é permitido excluir o usuário ADMIN padrão.")
                else:
                    excluir_usuario(user_id)
                    st.success("Usuário excluído com sucesso.")
                    st.rerun()

    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="section-card">', unsafe_allow_html=True)
    st.subheader("Configuração de e-mail")
    if all([SMTP_SERVER, SMTP_PORT, SMTP_USER, SMTP_PASSWORD, EMAIL_FROM]):
        st.success("SMTP configurado. Os e-mails automáticos estão habilitados.")
    else:
        st.warning("SMTP não configurado. Defina SMTP_SERVER, SMTP_PORT, SMTP_USER, SMTP_PASSWORD e EMAIL_FROM como variáveis de ambiente.")
    st.markdown('</div>', unsafe_allow_html=True)


# =========================================================
# PROJETOS CONCLUÍDOS
# =========================================================
@st.cache_data(ttl=60, show_spinner=False)
def listar_projetos_concluidos():
    conn = get_conn()
    try:
        df = pd.read_sql_query("""
            SELECT id, nome_projeto, tema, subtema, estado, municipio,
                   data_inicio, data_conclusao, prazo_real_meses,
                   custo_contratado, custo_final, observacoes, criado_por, criado_em
            FROM projetos_concluidos
            ORDER BY data_conclusao DESC NULLS LAST, id DESC
        """, conn)
    except Exception:
        df = pd.DataFrame()
    finally:
        conn.close()
    return df


def inserir_projeto_concluido(nome, tema, subtema, estado, municipio,
                               data_inicio, data_conclusao, custo_contratado,
                               custo_final, observacoes, criado_por):
    from datetime import date
    prazo_real = None
    if data_inicio and data_conclusao:
        try:
            di = data_inicio if isinstance(data_inicio, date) else date.fromisoformat(str(data_inicio))
            dc = data_conclusao if isinstance(data_conclusao, date) else date.fromisoformat(str(data_conclusao))
            delta_dias = (dc - di).days
            prazo_real = round(delta_dias / 30.44, 2)
        except Exception:
            prazo_real = None

    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO projetos_concluidos
            (nome_projeto, tema, subtema, estado, municipio,
             data_inicio, data_conclusao, prazo_real_meses,
             custo_contratado, custo_final, observacoes, criado_por, criado_em, atualizado_em)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        RETURNING id
    """, (nome, tema, subtema, estado, municipio,
          str(data_inicio) if data_inicio else None,
          str(data_conclusao) if data_conclusao else None,
          prazo_real, custo_contratado, custo_final,
          observacoes, criado_por, agora_str(), agora_str()))
    proj_id = cur.fetchone()[0]
    conn.commit()
    conn.close()
    return proj_id, prazo_real


def excluir_projeto_concluido(proj_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM projetos_concluidos WHERE id = %s", (proj_id,))
    conn.commit()
    conn.close()


def calcular_estatisticas_subtema(subtema: str):
    """Recalcula min/max/media de prazo_meses da view para o subtema."""
    conn = get_conn()
    try:
        df = pd.read_sql_query("""
            SELECT prazo_meses FROM vw_consulta_editais
            WHERE subtema = %s AND prazo_meses IS NOT NULL AND prazo_meses > 0
        """, conn, params=(subtema,))
    except Exception:
        df = pd.DataFrame()
    finally:
        conn.close()

    if df.empty:
        return None
    vals = df["prazo_meses"].tolist()
    s = sorted(vals)
    n = len(s)
    import math
    mean = sum(s)/n
    q1 = s[max(0, n//4)]
    q3 = s[min(n-1, 3*n//4)]
    iqr = q3 - q1
    return {
        "min": s[0], "max": s[-1], "mean": mean,
        "q1": q1, "q3": q3, "iqr": iqr,
        "lower": q1 - 1.5*iqr, "upper": q3 + 1.5*iqr,
        "n": n
    }


def pagina_projetos_concluidos():
    try:
        import plotly.graph_objects as go
        import plotly.express as px
        HAS_PLOTLY = True
    except ImportError:
        HAS_PLOTLY = False

    header_principal()
    st.markdown("## Projetos Concluídos")
    st.markdown("Registre projetos finalizados e compare o prazo real com as estimativas da análise estatística.")

    # ── Formulário de cadastro ──
    if st.session_state.perfil in ("ADMIN", "PMO"):
        with st.expander("Registrar novo projeto concluído", expanded=False):
            conn_view = get_conn()
            try:
                df_temas = pd.read_sql_query("SELECT DISTINCT tema, subtema FROM vw_consulta_editais WHERE tema IS NOT NULL ORDER BY tema, subtema", conn_view)
            except Exception:
                df_temas = pd.DataFrame(columns=["tema", "subtema"])
            finally:
                conn_view.close()

            # Seletores de tema/subtema FORA do form para permitir cascata dinâmica
            temas_disp = sorted(df_temas["tema"].dropna().unique().tolist())
            pc1, pc2 = st.columns(2)
            with pc1:
                tema_proj = st.selectbox("Tema", [""] + temas_disp, key="pc_form_tema")
            with pc2:
                subtemas_disp = sorted(df_temas[df_temas["tema"] == tema_proj]["subtema"].dropna().unique().tolist()) if tema_proj else []
                subtema_proj = st.selectbox("Subtema", [""] + subtemas_disp, key="pc_form_subtema")

            with st.form("form_proj_concluido", clear_on_submit=True):
                c1, c2 = st.columns(2)
                with c1:
                    nome_proj = st.text_input("Nome do projeto *")
                with c2:
                    estado_proj = st.text_input("Estado")

                c3, c4 = st.columns(2)
                with c3:
                    municipio_proj = st.text_input("Município")
                    data_inicio_proj = st.date_input("Data de início", value=None)
                with c4:
                    data_conclusao_proj = st.date_input("Data de conclusão", value=None)
                    custo_contratado_proj = st.number_input("Custo contratado (R$)", min_value=0.0, step=1000.0, format="%.2f")

                c5, c6 = st.columns(2)
                with c5:
                    custo_final_proj = st.number_input("Custo final realizado (R$)", min_value=0.0, step=1000.0, format="%.2f")
                with c6:
                    obs_proj = st.text_area("Observações", height=80)

                salvar = st.form_submit_button("Salvar projeto", type="primary")

            if salvar:
                if not nome_proj.strip():
                    st.warning("Informe o nome do projeto.")
                else:
                    proj_id, prazo_real = inserir_projeto_concluido(
                        nome=nome_proj.strip(), tema=tema_proj or None,
                        subtema=subtema_proj or None, estado=estado_proj or None,
                        municipio=municipio_proj or None,
                        data_inicio=data_inicio_proj, data_conclusao=data_conclusao_proj,
                        custo_contratado=custo_contratado_proj or None,
                        custo_final=custo_final_proj or None,
                        observacoes=obs_proj or None,
                        criado_por=st.session_state.usuario
                    )
                    prazo_msg = f" Prazo real calculado: **{prazo_real:.1f} meses**." if prazo_real else ""
                    st.success(f"Projeto registrado com sucesso!{prazo_msg}")
                    st.rerun()

    # ── Lista de projetos ──
    df_proj = listar_projetos_concluidos()

    if df_proj.empty:
        st.info("Nenhum projeto concluído registrado ainda.")
        return

    # ── Busca textual + Filtros ──
    busca_proj = st.text_input("🔍 Busca por nome do projeto", placeholder="Digite parte do nome...", key="pc_busca")

    cf1, cf2, cf3 = st.columns(3)
    with cf1:
        temas_f = ["Todos"] + sorted(df_proj["tema"].dropna().unique().tolist())
        tema_f = st.selectbox("Filtrar por tema", temas_f, key="pc_tema_f")
    with cf2:
        df_proj_f = df_proj[df_proj["tema"] == tema_f] if tema_f != "Todos" else df_proj
        subtemas_f = ["Todos"] + sorted(df_proj_f["subtema"].dropna().unique().tolist())
        subtema_f = st.selectbox("Filtrar por subtema", subtemas_f, key="pc_subtema_f")
    with cf3:
        estados_f = ["Todos"] + sorted(df_proj["estado"].dropna().unique().tolist())
        estado_f = st.selectbox("Filtrar por estado", estados_f, key="pc_estado_f")

    df_exib = df_proj.copy()
    if busca_proj.strip():
        df_exib = df_exib[df_exib["nome_projeto"].str.contains(busca_proj.strip(), case=False, na=False)]
    if tema_f != "Todos":
        df_exib = df_exib[df_exib["tema"] == tema_f]
    if subtema_f != "Todos":
        df_exib = df_exib[df_exib["subtema"] == subtema_f]
    if estado_f != "Todos":
        df_exib = df_exib[df_exib["estado"] == estado_f]

    # ── Métricas resumo ──
    m1, m2, m3, m4 = st.columns(4)
    prazo_vals = df_exib["prazo_real_meses"].dropna()
    custo_dif = (df_exib["custo_final"] - df_exib["custo_contratado"]).dropna()
    m1.metric("Projetos", len(df_exib))
    m2.metric("Prazo médio real", f"{prazo_vals.mean():.1f} m" if not prazo_vals.empty else "—")
    m3.metric("Prazo mínimo real", f"{prazo_vals.min():.1f} m" if not prazo_vals.empty else "—")
    m4.metric("Prazo máximo real", f"{prazo_vals.max():.1f} m" if not prazo_vals.empty else "—")

    # ── Tabela ──
    st.markdown("### Projetos registrados")
    colunas_exib = {
        "nome_projeto": "Projeto", "tema": "Tema", "subtema": "Subtema",
        "estado": "Estado", "municipio": "Município",
        "data_inicio": "Início", "data_conclusao": "Conclusão",
        "prazo_real_meses": "Prazo real (m)",
        "custo_contratado": "Custo contratado (R$)", "custo_final": "Custo final (R$)",
        "observacoes": "Obs.", "criado_por": "Registrado por"
    }
    df_tabela = df_exib[[c for c in colunas_exib if c in df_exib.columns]].rename(columns=colunas_exib)
    st.dataframe(df_tabela, use_container_width=True, hide_index=True)

    # Excluir projeto
    if st.session_state.perfil in ("ADMIN", "PMO"):
        with st.expander("Excluir projeto"):
            proj_id_del = st.selectbox(
                "Selecione o projeto",
                df_exib["id"].tolist(),
                format_func=lambda x: f"{x} — {df_exib.loc[df_exib['id']==x, 'nome_projeto'].values[0]}"
            )
            if st.button("Excluir projeto selecionado", type="primary"):
                excluir_projeto_concluido(proj_id_del)
                st.success("Projeto excluído.")
                st.rerun()

    # ── Comparação com Análise de Prazos ──
    st.markdown("---")
    st.markdown("### Comparação: Prazo Real vs. Estimativa Estatística")

    subtemas_comp = sorted(df_proj["subtema"].dropna().unique().tolist())
    if not subtemas_comp:
        st.info("Nenhum projeto com subtema definido para comparação.")
        return

    subtema_comp = st.selectbox("Selecione o subtema para comparar", subtemas_comp, key="pc_subtema_comp")
    df_comp = df_proj[df_proj["subtema"] == subtema_comp].dropna(subset=["prazo_real_meses"])

    est = calcular_estatisticas_subtema(subtema_comp)

    if df_comp.empty:
        st.info("Nenhum projeto concluído com prazo real para este subtema.")
    elif not est:
        st.info("Não há dados de editais suficientes para calcular a estimativa deste subtema.")
    else:
        # Métricas de comparação
        prazo_real_medio = df_comp["prazo_real_meses"].mean()
        est_min = est["lower"] if est["lower"] > 0 else est["min"]
        est_max = est["upper"] if est["upper"] > est["max"] else est["max"]
        est_medio = est["mean"]

        ce1, ce2, ce3, ce4 = st.columns(4)
        ce1.metric("Estimativa mínima (histórico)", f"{est['min']:.1f} m")
        ce2.metric("Estimativa máxima (histórico)", f"{est['max']:.1f} m")
        ce3.metric("Prazo real médio", f"{prazo_real_medio:.1f} m",
                   delta=f"{prazo_real_medio - est_medio:.1f} m vs. média estimada")
        dentro = df_comp[(df_comp["prazo_real_meses"] >= est["min"]) &
                         (df_comp["prazo_real_meses"] <= est["max"])]
        ce4.metric("Dentro do intervalo histórico", f"{len(dentro)}/{len(df_comp)}")

        if HAS_PLOTLY:
            # Gráfico 1: Barras comparando prazo real de cada projeto com intervalo estimado
            fig1 = go.Figure()

            # Faixa de intervalo histórico como área
            projetos_nomes = df_comp["nome_projeto"].tolist()
            prazos_reais = df_comp["prazo_real_meses"].tolist()

            fig1.add_trace(go.Bar(
                x=projetos_nomes, y=prazos_reais,
                name="Prazo real",
                marker_color=["#10b981" if est["min"] <= p <= est["max"] else "#ef4444" for p in prazos_reais],
                hovertemplate="<b>%{x}</b><br>Prazo real: %{y:.1f} meses<extra></extra>"
            ))

            # Linhas de referência
            fig1.add_hline(y=est["min"], line_dash="dash", line_color="#3b82f6",
                           annotation_text=f"Mín. histórico: {est['min']:.1f}m",
                           annotation_position="top right")
            fig1.add_hline(y=est["max"], line_dash="dash", line_color="#f59e0b",
                           annotation_text=f"Máx. histórico: {est['max']:.1f}m",
                           annotation_position="top right")
            fig1.add_hline(y=est["mean"], line_dash="dot", line_color="#8b5cf6",
                           annotation_text=f"Média histórica: {est['mean']:.1f}m",
                           annotation_position="top right")

            fig1.update_layout(
                title=f"Prazo real por projeto — {subtema_comp}",
                xaxis_title="Projeto", yaxis_title="Meses",
                height=400, template="plotly_white",
                showlegend=True
            )
            st.plotly_chart(fig1, use_container_width=True)
            st.caption("🟢 Dentro do intervalo histórico   🔴 Fora do intervalo histórico")

            # Gráfico 2: Dispersão prazo real ao longo do tempo
            df_comp_ord = df_comp.sort_values("data_conclusao")
            if not df_comp_ord["data_conclusao"].isna().all():
                fig2 = go.Figure()
                fig2.add_trace(go.Scatter(
                    x=df_comp_ord["data_conclusao"].astype(str).tolist(),
                    y=df_comp_ord["prazo_real_meses"].tolist(),
                    mode="markers+lines",
                    marker=dict(size=10, color="#2563eb"),
                    line=dict(color="#93c5fd", width=1, dash="dot"),
                    text=df_comp_ord["nome_projeto"].tolist(),
                    hovertemplate="<b>%{text}</b><br>Conclusão: %{x}<br>Prazo real: %{y:.1f} meses<extra></extra>",
                    name="Prazo real"
                ))
                fig2.add_hrect(y0=est["min"], y1=est["max"],
                               fillcolor="#3b82f6", opacity=0.08,
                               annotation_text="Intervalo histórico", annotation_position="top right")
                fig2.update_layout(
                    title="Evolução do prazo real ao longo do tempo",
                    xaxis_title="Data de conclusão", yaxis_title="Meses",
                    height=350, template="plotly_white"
                )
                st.plotly_chart(fig2, use_container_width=True)

            # Gráfico 3: Custo contratado vs realizado
            df_custo = df_comp.dropna(subset=["custo_contratado", "custo_final"])
            if not df_custo.empty:
                fig3 = go.Figure()
                fig3.add_trace(go.Bar(
                    name="Custo contratado",
                    x=df_custo["nome_projeto"].tolist(),
                    y=df_custo["custo_contratado"].tolist(),
                    marker_color="#3b82f6"
                ))
                fig3.add_trace(go.Bar(
                    name="Custo final realizado",
                    x=df_custo["nome_projeto"].tolist(),
                    y=df_custo["custo_final"].tolist(),
                    marker_color="#ef4444"
                ))
                fig3.update_layout(
                    title="Custo contratado vs. realizado",
                    xaxis_title="Projeto", yaxis_title="R$",
                    barmode="group", height=380, template="plotly_white"
                )
                st.plotly_chart(fig3, use_container_width=True)
        else:
            st.info("Instale plotly para ver os gráficos: pip install plotly")

    # Exportar
    st.markdown("---")
    csv_proj = df_exib.drop(columns=["id"], errors="ignore").to_csv(index=False).encode("utf-8")
    st.download_button("Exportar CSV", csv_proj, file_name="projetos_concluidos.csv", mime="text/csv")


# =========================================================
# ANÁLISE DE PRAZOS
# =========================================================
def exportar_analise_excel(data: dict) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    import math

    NAVY="FF0B1F3A"; BLUE="FF1A3F6F"; MID="FF1E5799"; ACCENT="FF1D6FC4"
    WHITE="FFFFFFFF"; LGRAY="FFF0F4F9"; MGRAY="FFE8EEF6"; DGRAY="FF3D5575"

    def hdr(ws,row,col,text,fg=WHITE,bg=NAVY,bold=True,size=10,merge_to=None,align="left"):
        c=ws.cell(row=row,column=col,value=text)
        c.font=Font(name="Arial",bold=bold,size=size,color=fg)
        c.fill=PatternFill("solid",fgColor=bg)
        c.alignment=Alignment(horizontal=align,vertical="center",wrap_text=True)
        if merge_to: ws.merge_cells(start_row=row,start_column=col,end_row=merge_to[0],end_column=merge_to[1])
        return c

    def val(ws,row,col,value,bold=False,size=10,color="FF0D1B2E",bg=None,align="left"):
        c=ws.cell(row=row,column=col,value=value)
        c.font=Font(name="Arial",bold=bold,size=size,color=color)
        if bg: c.fill=PatternFill("solid",fgColor=bg)
        c.alignment=Alignment(horizontal=align,vertical="center",wrap_text=True)
        return c

    def sec(ws,row,col,text,cols=2):
        ws.merge_cells(start_row=row,start_column=col,end_row=row,end_column=col+cols-1)
        c=ws.cell(row=row,column=col,value=text)
        c.font=Font(name="Arial",bold=True,size=10,color=WHITE)
        c.fill=PatternFill("solid",fgColor=BLUE)
        c.alignment=Alignment(horizontal="left",vertical="center")
        ws.row_dimensions[row].height=18
        return c

    def borders(ws,r1,r2,c1,c2):
        s=Side(style="thin",color="FFD8E5F2")
        b=Border(left=s,right=s,top=s,bottom=s)
        for r in range(r1,r2+1):
            for c in range(c1,c2+1):
                ws.cell(r,c).border=b

    def cols_w(ws,d):
        for c,w in d.items(): ws.column_dimensions[get_column_letter(c)].width=w

    wb=openpyxl.Workbook()

    # ─── ABA 1: Resumo ───────────────────────────────────
    ws1=wb.active; ws1.title="Resumo da Análise"
    ws1.sheet_view.showGridLines=False
    cols_w(ws1,{1:34,2:30})

    ws1.merge_cells("A1:B1")
    c=ws1.cell(1,1,"RELATÓRIO DE ANÁLISE DE PROJETOS")
    c.font=Font(name="Arial",bold=True,size=14,color=WHITE)
    c.fill=PatternFill("solid",fgColor=NAVY)
    c.alignment=Alignment(horizontal="center",vertical="center")
    ws1.row_dimensions[1].height=36

    from datetime import datetime
    ws1.merge_cells("A2:B2")
    c2=ws1.cell(2,1,f"Gerado em: {datetime.now().strftime('%d/%m/%Y, %H:%M:%S')}")
    c2.font=Font(name="Arial",size=9,color=DGRAY)
    c2.fill=PatternFill("solid",fgColor=MGRAY)
    c2.alignment=Alignment(horizontal="center",vertical="center")

    r=4
    sec(ws1,r,1,"PARÂMETROS DA ANÁLISE",2); r+=1
    for label,value in [
        ("Tema",data.get("tema","")),("Subtema",data.get("subtema","")),
        ("País",data.get("pais_sel","Todos")),("Estado",data.get("estado_sel","Todos")),
        ("Tema SAN","Sim" if data.get("is_san") else "Não"),
    ]:
        val(ws1,r,1,label,bold=True,color=DGRAY,bg=LGRAY); val(ws1,r,2,value); r+=1
    borders(ws1,5,r-1,1,2)

    pearson=data.get("pearson",0); spearman=data.get("spearman",0)
    corr_forte=data.get("corr_forte",False); reg=data.get("reg"); reg_label=data.get("reg_label","")

    def forca(v):
        a=abs(v)
        if a>=0.7: return "Forte"
        if a>=0.6: return "Moderada"
        if a>=0.3: return "Fraca"
        return "Muito fraca"

    r+=1; sec(ws1,r,1,"CORRELAÇÃO",2); r+=1
    for label,value in [
        ("Pearson (Linear)",f"{pearson:.4f}"),("Spearman (Não-linear)",f"{spearman:.4f}"),
        ("Correlação Dominante","Logarítmica (Spearman)" if abs(spearman)>abs(pearson) else "Linear (Pearson)"),
        ("Status da Correlação",f"{'FORTE (≥ 0,6) — usando regressão' if corr_forte else 'FRACA (< 0,6) — usando histórico'}"),
        ("Tipo de Regressão Utilizado",data.get("reg_type_display","Automático")),
    ]:
        val(ws1,r,1,label,bold=True,color=DGRAY,bg=LGRAY); val(ws1,r,2,value); r+=1
    borders(ws1,r-5,r-1,1,2)

    if reg:
        r+=1; sec(ws1,r,1,"REGRESSÃO",2); r+=1
        for label,value in [
            ("Tipo",reg.get("type","").capitalize()),("Equação",reg.get("eq","")),
            ("R²",f"{reg.get('r2',0):.4f}"),("1 - R²",f"{1-reg.get('r2',0):.4f}"),
            ("Coeficiente a",f"{reg.get('a',0):.4f}"),("Coeficiente b",f"{reg.get('b',0):.4f}"),
        ]:
            val(ws1,r,1,label,bold=True,color=DGRAY,bg=LGRAY); val(ws1,r,2,value); r+=1
        borders(ws1,r-6,r-1,1,2)

    st_p=data.get("st_prazos",{})
    r+=1; sec(ws1,r,1,"INTERVALO HISTÓRICO OBSERVADO",2); r+=1
    val(ws1,r,1,"Prazo Mínimo Histórico (meses)",bold=True,color=DGRAY,bg=LGRAY)
    val(ws1,r,2,f"{st_p.get('min',0):.2f}"); r+=1
    val(ws1,r,1,"Prazo Máximo Histórico (meses)",bold=True,color=DGRAY,bg=LGRAY)
    val(ws1,r,2,f"{st_p.get('max',0):.2f}"); r+=1
    borders(ws1,r-2,r-1,1,2)

    alerts=data.get("alerts",[])
    if alerts:
        r+=1
        ws1.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
        c=ws1.cell(r,1,"⚠ ALERTA: ESTIMATIVA FORA DO INTERVALO HISTÓRICO")
        c.font=Font(name="Arial",bold=True,size=10,color=WHITE)
        c.fill=PatternFill("solid",fgColor="FFDC2626")
        c.alignment=Alignment(horizontal="left",vertical="center")
        ws1.row_dimensions[r].height=18; r+=1
        for a in alerts:
            ws1.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
            c=ws1.cell(r,1,a)
            c.font=Font(name="Arial",size=9,color="FFDC2626")
            c.fill=PatternFill("solid",fgColor="FFFFF1F1")
            c.alignment=Alignment(wrap_text=True); r+=1

    # ─── ABA 2: Cronograma ───────────────────────────────
    ws2=wb.create_sheet("Cronograma Kerzner")
    ws2.sheet_view.showGridLines=False
    cols_w(ws2,{1:30,2:22,3:22,4:22})
    ws2.merge_cells("A1:D1")
    c=ws2.cell(1,1,"CRONOGRAMA KERZNER (2009)")
    c.font=Font(name="Arial",bold=True,size=13,color=WHITE)
    c.fill=PatternFill("solid",fgColor=NAVY)
    c.alignment=Alignment(horizontal="center",vertical="center")
    ws2.row_dimensions[1].height=32

    r=3; kr=data.get("kerzner",{})
    enc_desc="10% da execução (SAN)" if data.get("is_san") else "10% do total (mín. 1 mês)"
    for label,value in [
        ("Metodologia","Kerzner (2009)"),("Encerramento",enc_desc),
        ("Método de Cálculo",f"Regressão ({reg_label})" if corr_forte else "Valores históricos"),
        ("Equação utilizada",reg.get("eq","—") if reg else "—"),
        ("R²",f"{reg.get('r2',0):.4f}" if reg else "—"),
    ]:
        val(ws2,r,1,label,bold=True,color=DGRAY,bg=LGRAY)
        ws2.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4)
        val(ws2,r,2,value); r+=1
    borders(ws2,3,r-1,1,4)

    r+=1
    for i,h in enumerate(["FASE","PERCENTUAL","CENÁRIO MÍNIMO (meses)","CENÁRIO MÁXIMO (meses)"]):
        hdr(ws2,r,i+1,h,bg=NAVY,align="center")
    ws2.row_dimensions[r].height=20; r+=1

    total_min=kr.get("total_min",0); total_max=kr.get("total_max",0)
    for i,(fase,pct,vmin,vmax) in enumerate([
        ("Planejamento","50% do total",kr.get("plan_min",0),kr.get("plan_max",0)),
        ("Execução","40% do total",kr.get("exec_min",0),kr.get("exec_max",0)),
        ("Encerramento",enc_desc,kr.get("enc_min",0),kr.get("enc_max",0)),
        ("TOTAL DO PROJETO","100%",total_min,total_max),
    ]):
        bg=LGRAY if i%2==0 else WHITE
        bold=fase=="TOTAL DO PROJETO"
        if bold: bg=MGRAY
        val(ws2,r,1,fase,bold=bold,bg=bg)
        val(ws2,r,2,pct,align="center",bg=bg)
        val(ws2,r,3,f"{vmin:.4f}",align="center",bg=bg,bold=bold)
        val(ws2,r,4,f"{vmax:.4f}",align="center",bg=bg,bold=bold); r+=1
    borders(ws2,r-4,r-1,1,4)

    r+=1
    ws2.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
    c=ws2.cell(r,1,"DETALHAMENTO DO PLANEJAMENTO (Kerzner)")
    c.font=Font(name="Arial",bold=True,size=10,color=WHITE)
    c.fill=PatternFill("solid",fgColor=MID)
    c.alignment=Alignment(horizontal="left",vertical="center")
    ws2.row_dimensions[r].height=18; r+=1
    for i,h in enumerate(["Sub-fase","Percentual do Total","Cenário Mínimo (meses)","Cenário Máximo (meses)"]):
        hdr(ws2,r,i+1,h,bg=MID,align="center")
    ws2.row_dimensions[r].height=18; r+=1
    for i,(nome,pct) in enumerate([("Conceitualização",0.05),("Estudo de Viabilidade",0.10),
                                    ("Planejamento Preliminar",0.15),("Planejamento Detalhado",0.20)]):
        bg=LGRAY if i%2==0 else WHITE
        val(ws2,r,1,nome,bg=bg)
        val(ws2,r,2,f"{int(pct*100)}%",align="center",bg=bg)
        val(ws2,r,3,f"{total_min*pct:.4f}",align="center",bg=bg)
        val(ws2,r,4,f"{total_max*pct:.4f}",align="center",bg=bg); r+=1
    borders(ws2,r-4,r-1,1,4)

    # ─── ABA 3: Estatísticas ─────────────────────────────
    ws3=wb.create_sheet("Estatísticas")
    ws3.sheet_view.showGridLines=False
    cols_w(ws3,{1:28,2:18,3:18})
    ws3.merge_cells("A1:C1")
    c=ws3.cell(1,1,"ESTATÍSTICAS DA ANÁLISE")
    c.font=Font(name="Arial",bold=True,size=13,color=WHITE)
    c.fill=PatternFill("solid",fgColor=NAVY)
    c.alignment=Alignment(horizontal="center",vertical="center")
    ws3.row_dimensions[1].height=32

    r=3
    for i,h in enumerate(["INDICADOR","ESFORÇO","PRAZO (meses)"]):
        hdr(ws3,r,i+1,h,bg=NAVY,align="center")
    ws3.row_dimensions[r].height=20; r+=1

    st_e=data.get("st_esforcos",{})
    for i,(label,key) in enumerate([
        ("Média","mean"),("Mediana","median"),("Desvio Padrão","std"),
        ("1º Quartil (Q1)","q1"),("3º Quartil (Q3)","q3"),("IQR (Q3 - Q1)","iqr"),
        ("Limite Inferior","lower"),("Limite Superior","upper"),
    ]):
        bg=LGRAY if i%2==0 else WHITE
        val(ws3,r,1,label,bold=True,color=DGRAY,bg=bg)
        val(ws3,r,2,f"{st_e.get(key,0):.2f}",align="center",bg=bg)
        val(ws3,r,3,f"{st_p.get(key,0):.2f}",align="center",bg=bg); r+=1
    borders(ws3,4,r-1,1,3)

    r+=1
    for i,h in enumerate(["CORRELAÇÃO","VALOR","FORÇA"]):
        hdr(ws3,r,i+1,h,bg=NAVY,align="center")
    ws3.row_dimensions[r].height=20; r+=1
    val(ws3,r,1,"Pearson (Linear)",bold=True,color=DGRAY,bg=LGRAY)
    val(ws3,r,2,f"{pearson:.4f}",align="center",bg=LGRAY)
    val(ws3,r,3,forca(pearson),align="center",bg=LGRAY); r+=1
    val(ws3,r,1,"Spearman (Não-linear)",bold=True,color=DGRAY)
    val(ws3,r,2,f"{spearman:.4f}",align="center")
    val(ws3,r,3,forca(spearman),align="center"); r+=1
    borders(ws3,r-2,r-1,1,3)

    # ─── ABA 4: Projetos ─────────────────────────────────
    ws4=wb.create_sheet("Projetos")
    ws4.sheet_view.showGridLines=False
    cols_w(ws4,{1:55,2:10,3:16,4:12,5:12,6:14,7:14,8:16})
    ws4.merge_cells("A1:H1")
    c=ws4.cell(1,1,"DETALHAMENTO DOS PROJETOS")
    c.font=Font(name="Arial",bold=True,size=13,color=WHITE)
    c.fill=PatternFill("solid",fgColor=NAVY)
    c.alignment=Alignment(horizontal="center",vertical="center")
    ws4.row_dimensions[1].height=32

    r=3
    for i,h in enumerate(["Nome","País","Estado","Tipo","Esforço","Duração (meses)","Classificação","Status na Análise"]):
        hdr(ws4,r,i+1,h,bg=NAVY,align="center")
    ws4.row_dimensions[r].height=20; r+=1

    projetos=data.get("projetos",[])
    for i,p in enumerate(projetos):
        bg=LGRAY if i%2==0 else WHITE
        classif=p.get("Classificação","Normal")
        if classif=="Outlier Alto": bg="FFFFF1F1"
        elif classif=="Outlier Baixo": bg="FFF0F9FF"
        status="Excluído" if p.get("_excluido") else "Incluído"
        for j,v in enumerate([
            p.get("nome_edital",""),p.get("pais",""),p.get("estado",""),
            "Esforço",p.get("esforco",""),p.get("prazo_meses",""),classif,status
        ]):
            val(ws4,r,j+1,v,bg=bg,align="left" if j==0 else "center")
        ws4.row_dimensions[r].height=40; r+=1
    borders(ws4,4,r-1,1,8)

    r+=1
    total=len(projetos); analisados=len([p for p in projetos if not p.get("_excluido")])
    ws4.cell(r,4,f"Total: {total} projetos").font=Font(name="Arial",bold=True,size=9)
    ws4.cell(r,5,f"Analisados: {analisados}").font=Font(name="Arial",bold=True,size=9)
    ws4.cell(r,6,f"Excluídos: {total-analisados}").font=Font(name="Arial",bold=True,size=9)

    buf=BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


def calcular_pearson(x, y):
    import math
    n = len(x)
    if n < 2:
        return 0
    mx, my = sum(x)/n, sum(y)/n
    num = sum((x[i]-mx)*(y[i]-my) for i in range(n))
    dx = math.sqrt(sum((v-mx)**2 for v in x))
    dy = math.sqrt(sum((v-my)**2 for v in y))
    if dx == 0 or dy == 0:
        return 0
    return num / (dx * dy)


def calcular_spearman(x, y):
    n = len(x)
    if n < 2:
        return 0

    def rank_com_empates(vals):
        # Atribui média dos ranks para valores empatados (tie-average)
        sorted_vals = sorted(enumerate(vals), key=lambda t: t[1])
        ranks = [0.0] * n
        i = 0
        while i < n:
            j = i
            # Encontra grupo de empate
            while j < n - 1 and sorted_vals[j][1] == sorted_vals[j+1][1]:
                j += 1
            # Rank médio para o grupo (1-indexed)
            rank_medio = (i + j) / 2 + 1
            for k in range(i, j + 1):
                ranks[sorted_vals[k][0]] = rank_medio
            i = j + 1
        return ranks

    rx = rank_com_empates(x)
    ry = rank_com_empates(y)
    return calcular_pearson(rx, ry)


def regressao_linear(x, y):
    import math
    n = len(x)
    sx, sy = sum(x), sum(y)
    sxy = sum(x[i]*y[i] for i in range(n))
    sx2 = sum(v**2 for v in x)
    denom = n*sx2 - sx**2
    if abs(denom) < 1e-10:
        return None
    a = (n*sxy - sx*sy) / denom
    b = (sy - a*sx) / n
    my = sy/n
    ss_tot = sum((v-my)**2 for v in y)
    ss_res = sum((y[i] - (a*x[i]+b))**2 for i in range(n))
    r2 = 1 - ss_res/ss_tot if ss_tot > 0 else 0
    return {"type": "linear", "a": a, "b": b, "r2": r2,
            "eq": f"y = {a:.4f} x + {b:.4f}",
            "predict": lambda xi, _a=a, _b=b: _a*xi + _b}


def regressao_logaritmica(x, y):
    import math
    lx = [math.log(max(v, 0.001)) for v in x]
    result = regressao_linear(lx, y)
    if result:
        a, b = result["a"], result["b"]
        result["eq"] = f"y = {a:.4f} * ln(x) + {b:.4f}"
        result["type"] = "logaritmica"
        result["predict"] = lambda xi, _a=a, _b=b: _a * __import__("math").log(max(xi, 0.001)) + _b
    return result




from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO


# ── Paleta de cores FGV ──────────────────────────────────────────
C_NAVY   = "0B1F3A"   # cabeçalhos principais
C_BLUE   = "1A3F6F"   # cabeçalhos de seção
C_MID    = "1E5799"   # cabeçalhos de tabela
C_LIGHT  = "D6E4F0"   # fundo de linhas de dados
C_WHITE  = "FFFFFF"
C_WARN   = "FFF3CD"   # fundo alerta
C_WARN_B = "FF8C00"   # borda alerta
C_NORM   = "D4EDDA"   # normal
C_LOW    = "CCE5FF"   # outlier baixo
C_HIGH   = "F8D7DA"   # outlier alto
C_EXCL   = "E2E3E5"   # excluído

def _ft(bold=False, color=C_WHITE, size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border(style="thin"):
    s = Side(border_style=style, color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _header_row(ws, row, text, span_end_col, bg=C_NAVY, fg=C_WHITE, size=12):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = _ft(bold=True, color=fg, size=size)
    cell.fill = _fill(bg)
    cell.alignment = _align("left")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_end_col)


def _section(ws, row, text, span_end_col):
    _header_row(ws, row, text, span_end_col, bg=C_BLUE, size=10)


def _kv(ws, row, label, value, label_color="334155", value_color="0D1B2E"):
    lc = ws.cell(row=row, column=1, value=label)
    lc.font = _ft(bold=True, color=label_color, size=10)
    lc.alignment = _align()
    lc.fill = _fill("F8FAFC")
    vc = ws.cell(row=row, column=2, value=value)
    vc.font = _ft(color=value_color, size=10)
    vc.alignment = _align()


def gerar_relatorio_excel(
    subtema, tema, pais_sel, estado_sel,
    pearson, spearman, reg, reg_type_label,
    st_prazos, st_esforcos,
    exec_min, exec_max, plan_min, plan_max,
    enc_min, enc_max, total_min, total_max,
    metodo, is_san, hist_exec_min, hist_exec_max,
    df_class, excluir_ids,
) -> bytes:
    wb = Workbook()
    now_str = datetime.now().strftime("%d/%m/%Y, %H:%M:%S")
    max_corr = max(abs(pearson), abs(spearman))
    corr_forte = max_corr >= 0.6

    def corr_forca(v):
        a = abs(v)
        if a >= 0.7: return "Forte"
        if a >= 0.6: return "Moderada"
        if a >= 0.3: return "Fraca"
        return "Muito fraca"

    # ── ABA 1: Resumo da Análise ─────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumo da Análise"
    ws1.column_dimensions["A"].width = 38
    ws1.column_dimensions["B"].width = 42

    _header_row(ws1, 1, "RELATÓRIO DE ANÁLISE DE PROJETOS", 2, size=13)
    ws1.cell(row=2, column=1, value=f"Gerado em: {now_str}").font = _ft(color="64748B", size=9, italic=True)
    ws1.merge_cells("A2:B2")

    ws1.row_dimensions[3].height = 6

    _section(ws1, 4, "PARÂMETROS DA ANÁLISE", 2)
    _kv(ws1, 5, "Tema", tema or "Todos")
    _kv(ws1, 6, "Subtema", subtema)
    _kv(ws1, 7, "País", pais_sel)
    _kv(ws1, 8, "Estado", estado_sel)
    _kv(ws1, 9, "Tema SAN", "Sim" if is_san else "Não")

    ws1.row_dimensions[10].height = 6
    _section(ws1, 11, "CORRELAÇÃO", 2)
    _kv(ws1, 12, "Pearson (Linear)", f"{pearson:.4f}  —  {corr_forca(pearson)}")
    _kv(ws1, 13, "Spearman (Não-linear)", f"{spearman:.4f}  —  {corr_forca(spearman)}")
    dom = "Linear (Pearson)" if abs(pearson) >= abs(spearman) else "Logarítmica (Spearman)"
    _kv(ws1, 14, "Correlação Dominante", dom)
    status_corr = f"FORTE (≥ 0,6) — usando regressão" if corr_forte else f"FRACA (< 0,6) — usando histórico"
    _kv(ws1, 15, "Status da Correlação", status_corr)
    _kv(ws1, 16, "Tipo de Regressão Utilizado", reg_type_label if reg else "—")

    row = 17
    if reg and corr_forte:
        ws1.row_dimensions[row].height = 6; row += 1
        _section(ws1, row, "REGRESSÃO", 2); row += 1
        tipo_r = "Linear" if reg.get("type") == "linear" else "Logarítmica"
        _kv(ws1, row, "Tipo", tipo_r); row += 1
        _kv(ws1, row, "Equação", reg["eq"]); row += 1
        _kv(ws1, row, "R²", f"{reg['r2']:.4f}"); row += 1
        _kv(ws1, row, "1 - R²", f"{1-reg['r2']:.4f}"); row += 1
        _kv(ws1, row, "Coeficiente a", f"{reg['a']:.4f}"); row += 1
        _kv(ws1, row, "Coeficiente b", f"{reg['b']:.4f}"); row += 1

    ws1.row_dimensions[row].height = 6; row += 1
    _section(ws1, row, "INTERVALO HISTÓRICO DE EXECUÇÃO OBSERVADO", 2); row += 1
    _kv(ws1, row, "Prazo Mínimo Histórico de Execução (meses)", f"{hist_exec_min:.2f}"); row += 1
    _kv(ws1, row, "Prazo Máximo Histórico de Execução (meses)", f"{hist_exec_max:.2f}"); row += 1

    # Alertas — só execução
    alerts = []
    if exec_min > hist_exec_max:
        alerts.append(f"Prazo mínimo de execução estimado ({exec_min:.2f} m) ACIMA do máximo histórico ({hist_exec_max:.2f} m)")
    if exec_max > hist_exec_max:
        alerts.append(f"Prazo máximo de execução estimado ({exec_max:.2f} m) ACIMA do máximo histórico ({hist_exec_max:.2f} m)")

    if alerts:
        ws1.row_dimensions[row].height = 6; row += 1
        c = ws1.cell(row=row, column=1, value="⚠ ALERTA: ESTIMATIVA DE EXECUÇÃO FORA DO INTERVALO HISTÓRICO")
        c.font = _ft(bold=True, color="92400E", size=10)
        c.fill = _fill(C_WARN)
        c.alignment = _align()
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 1
        for a in alerts:
            ac = ws1.cell(row=row, column=2, value=a)
            ac.font = _ft(color="92400E", size=9)
            ac.fill = _fill(C_WARN)
            row += 1

    # ── ABA 2: Cronograma Kerzner ─────────────────────────────────
    ws2 = wb.create_sheet("Cronograma Kerzner")
    for col, w in enumerate([28, 28, 22, 22], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    _header_row(ws2, 1, "CRONOGRAMA KERZNER (2009)", 4, size=13)
    ws2.row_dimensions[2].height = 6

    enc_desc = "10% do tempo de execução (tema SAN)" if is_san else "10% do total (mín. 1 mês)"
    metas = [
        ("Metodologia", "Kerzner (2009)"),
        ("Encerramento", enc_desc),
        ("Método de Cálculo", metodo),
    ]
    if reg and corr_forte:
        metas += [("Equação utilizada", reg["eq"]), ("R²", f"{reg['r2']:.4f}")]

    for i, (k, v) in enumerate(metas, 3):
        _kv(ws2, i, k, v)

    r = 3 + len(metas) + 1
    ws2.row_dimensions[r].height = 6; r += 1

    # Header tabela fases
    for col, txt in enumerate(["FASE", "PERCENTUAL", "CENÁRIO MÍNIMO (meses)", "CENÁRIO MÁXIMO (meses)"], 1):
        c = ws2.cell(row=r, column=col, value=txt)
        c.font = _ft(bold=True, color=C_WHITE, size=10)
        c.fill = _fill(C_MID)
        c.alignment = _align("center")
    r += 1

    fases = [
        ("Planejamento", "50% do total", plan_min, plan_max),
        ("Execução",     "40% do total", exec_min, exec_max),
        ("Encerramento", "10% do total" if not is_san else "10% da execução", enc_min, enc_max),
        ("TOTAL DO PROJETO", "100%", total_min, total_max),
    ]
    for i, (nome, pct, vmin, vmax) in enumerate(fases):
        bg = C_LIGHT if i % 2 == 0 else C_WHITE
        bold = nome == "TOTAL DO PROJETO"
        for col, val in enumerate([nome, pct, round(vmin,4), round(vmax,4)], 1):
            c = ws2.cell(row=r, column=col, value=val)
            c.font = _ft(bold=bold, color="0D1B2E", size=10)
            c.fill = _fill(C_NAVY if bold else bg)
            if bold: c.font = _ft(bold=True, color=C_WHITE, size=10)
            c.alignment = _align("center" if col > 1 else "left")
        r += 1

    ws2.row_dimensions[r].height = 8; r += 1
    _section(ws2, r, "DETALHAMENTO DO PLANEJAMENTO (Kerzner)", 4); r += 1
    for col, txt in enumerate(["Sub-fase", "Percentual do Total", "Cenário Mínimo (meses)", "Cenário Máximo (meses)"], 1):
        c = ws2.cell(row=r, column=col, value=txt)
        c.font = _ft(bold=True, color=C_WHITE, size=9)
        c.fill = _fill(C_BLUE)
        c.alignment = _align("center")
    r += 1

    subfases = [
        ("Conceitualização",      0.05),
        ("Estudo de Viabilidade", 0.10),
        ("Planejamento Preliminar", 0.15),
        ("Planejamento Detalhado", 0.20),
    ]
    for i, (nome, pct) in enumerate(subfases):
        bg = C_LIGHT if i % 2 == 0 else C_WHITE
        for col, val in enumerate([nome, f"{int(pct*100)}%",
                                    round(total_min*pct, 4), round(total_max*pct, 4)], 1):
            c = ws2.cell(row=r, column=col, value=val)
            c.font = _ft(color="0D1B2E", size=10)
            c.fill = _fill(bg)
            c.alignment = _align("center" if col > 1 else "left")
        r += 1

    # ── ABA 3: Estatísticas ───────────────────────────────────────
    ws3 = wb.create_sheet("Estatísticas")
    for col, w in enumerate([30, 20, 20], 1):
        ws3.column_dimensions[get_column_letter(col)].width = w

    _header_row(ws3, 1, "ESTATÍSTICAS DA ANÁLISE", 3, size=13)
    ws3.row_dimensions[2].height = 6

    for col, txt in enumerate(["INDICADOR", "ESFORÇO", "PRAZO (meses)"], 1):
        c = ws3.cell(row=3, column=col, value=txt)
        c.font = _ft(bold=True, color=C_WHITE, size=10)
        c.fill = _fill(C_MID)
        c.alignment = _align("center")

    indicadores = [
        ("Média",           st_esforcos.get("mean",0),   st_prazos.get("mean",0)),
        ("Mediana",         st_esforcos.get("median",0), st_prazos.get("median",0)),
        ("Desvio Padrão",   st_esforcos.get("std",0),    st_prazos.get("std",0)),
        ("1º Quartil (Q1)", st_esforcos.get("q1",0),     st_prazos.get("q1",0)),
        ("3º Quartil (Q3)", st_esforcos.get("q3",0),     st_prazos.get("q3",0)),
        ("IQR (Q3 - Q1)",   st_esforcos.get("iqr",0),    st_prazos.get("iqr",0)),
        ("Limite Inferior", st_esforcos.get("lower",0),  st_prazos.get("lower",0)),
        ("Limite Superior", st_esforcos.get("upper",0),  st_prazos.get("upper",0)),
        ("Mínimo histórico",st_esforcos.get("min",0),    st_prazos.get("min",0)),
        ("Máximo histórico",st_esforcos.get("max",0),    st_prazos.get("max",0)),
    ]
    for i, (label, esf, prazo) in enumerate(indicadores, 4):
        bg = C_LIGHT if i % 2 == 0 else C_WHITE
        for col, val in enumerate([label, round(esf,4), round(prazo,4)], 1):
            c = ws3.cell(row=i, column=col, value=val)
            c.font = _ft(color="0D1B2E", size=10)
            c.fill = _fill(bg)
            c.alignment = _align("center" if col > 1 else "left")

    r3 = len(indicadores) + 5
    ws3.row_dimensions[r3].height = 8; r3 += 1
    _section(ws3, r3, "CORRELAÇÃO", 3); r3 += 1
    for col, txt in enumerate(["CORRELAÇÃO", "VALOR", "FORÇA"], 1):
        c = ws3.cell(row=r3, column=col, value=txt)
        c.font = _ft(bold=True, color=C_WHITE, size=10)
        c.fill = _fill(C_BLUE)
        c.alignment = _align("center")
    r3 += 1
    for i, (label, val) in enumerate([("Pearson (Linear)", pearson), ("Spearman (Não-linear)", spearman)]):
        bg = C_LIGHT if i % 2 == 0 else C_WHITE
        for col, v in enumerate([label, round(val, 4), corr_forca(val)], 1):
            c = ws3.cell(row=r3, column=col, value=v)
            c.font = _ft(color="0D1B2E", size=10)
            c.fill = _fill(bg)
            c.alignment = _align("center" if col > 1 else "left")
        r3 += 1

    # ── ABA 4: Projetos ───────────────────────────────────────────
    ws4 = wb.create_sheet("Projetos")
    ws4.column_dimensions["A"].width = 55
    for col, w in enumerate([14, 18, 12, 12, 18, 18, 16], 2):
        ws4.column_dimensions[get_column_letter(col)].width = w

    _header_row(ws4, 1, "DETALHAMENTO DOS PROJETOS", 8, size=13)
    ws4.row_dimensions[2].height = 6

    headers = ["Nome", "País", "Estado", "Tipo", "Esforço", "Duração (meses)", "Classificação", "Status na Análise"]
    for col, txt in enumerate(headers, 1):
        c = ws4.cell(row=3, column=col, value=txt)
        c.font = _ft(bold=True, color=C_WHITE, size=10)
        c.fill = _fill(C_MID)
        c.alignment = _align("center")

    row4 = 4
    total = len(df_class)
    excluidos_count = len(excluir_ids)
    analisados = total - excluidos_count

    for _, proj in df_class.iterrows():
        excluido = proj["id"] in excluir_ids
        classif = proj.get("Classificação", "Normal")
        status = "Excluído" if excluido else "Incluído"

        if excluido:
            bg = C_EXCL
        elif classif == "Outlier Alto":
            bg = C_HIGH
        elif classif == "Outlier Baixo":
            bg = C_LOW
        else:
            bg = C_NORM if row4 % 2 == 0 else C_WHITE

        vals = [
            str(proj.get("nome_edital") or "—"),
            str(proj.get("pais") or "—"),
            str(proj.get("estado") or "—"),
            str(proj.get("tipo_edital") or "Esforço"),
            proj.get("esforco") or 0,
            abs(proj.get("prazo_meses") or 0),
            classif,
            status,
        ]
        for col, val in enumerate(vals, 1):
            c = ws4.cell(row=row4, column=col, value=val)
            c.font = _ft(color="0D1B2E", size=9)
            c.fill = _fill(bg)
            c.alignment = _align("left" if col == 1 else "center", wrap=col==1)
        row4 += 1

    # Rodapé totais
    ws4.row_dimensions[row4].height = 6; row4 += 1
    summary = f"Total: {total} projetos   |   Analisados: {analisados}   |   Excluídos: {excluidos_count}"
    c = ws4.cell(row=row4, column=1, value=summary)
    c.font = _ft(bold=True, color="1A3F6F", size=10)
    ws4.merge_cells(start_row=row4, start_column=1, end_row=row4, end_column=8)

    # ── Salvar ────────────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def pagina_analise_prazos():
    import math

    header_principal()
    st.markdown("## Análise de Prazos por Tema")

    df_edit = carregar_view()
    if df_edit.empty:
        st.warning("Nenhum dado disponível.")
        return

    for col in ["prazo_meses", "esforco"]:
        if col in df_edit.columns:
            df_edit[col] = pd.to_numeric(df_edit[col], errors="coerce")

    # ── Filtros ──
    cf1, cf2, cf3, cf4 = st.columns(4)
    with cf1:
        temas = sorted(df_edit["tema"].dropna().unique().tolist())
        tema_sel = st.selectbox("Tema", ["Todos"] + temas, key="ap_tema")
    df_filt = df_edit if tema_sel == "Todos" else df_edit[df_edit["tema"] == tema_sel]
    with cf2:
        subtemas = sorted(df_filt["subtema"].dropna().unique().tolist())
        subtema_sel = st.selectbox("Subtema (opcional)", ["Todos"] + subtemas, key="ap_subtema")
    if tema_sel == "Todos" and subtema_sel == "Todos":
        st.info("Selecione ao menos um Tema ou Subtema para iniciar a análise.")
        return
    if subtema_sel != "Todos":
        df_tema = df_filt[df_filt["subtema"] == subtema_sel].copy()
    else:
        df_tema = df_filt.copy()
    with cf3:
        paises = sorted(df_tema["pais"].dropna().unique().tolist())
        pais_sel = st.selectbox("País", ["Todos"] + paises, key="ap_pais")
    if pais_sel != "Todos":
        df_tema = df_tema[df_tema["pais"] == pais_sel]
    with cf4:
        estados = sorted(df_tema["estado"].dropna().unique().tolist())
        estado_sel = st.selectbox("Estado", ["Todos"] + estados, key="ap_estado")
    if estado_sel != "Todos":
        df_tema = df_tema[df_tema["estado"] == estado_sel]

    # ── Exclusões ──
    label_analise = subtema_sel if subtema_sel != "Todos" else tema_sel
    excluir_key = f"excluir_{label_analise}_{pais_sel}_{estado_sel}"
    if excluir_key not in st.session_state:
        st.session_state[excluir_key] = set()
    df_tema["_excluido"] = df_tema["id"].isin(st.session_state[excluir_key])
    df_analise = df_tema[~df_tema["_excluido"]].copy()

    if len(df_analise) < 3:
        st.warning(f"São necessários pelo menos 3 projetos. Encontrados: {len(df_analise)}")
        return

    df_valido = df_analise.dropna(subset=["prazo_meses", "esforco"])
    df_valido = df_valido[(df_valido["prazo_meses"] > 0) & (df_valido["esforco"] > 0)]
    is_san = "SAN-" in subtema_sel.upper() or (subtema_sel == "Todos" and "SAN-" in tema_sel.upper())

    # ── Estatísticas ──
    def stats(vals):
        if not vals: return {}
        s = sorted(vals)
        n = len(s)
        mean = sum(s)/n
        median = s[n//2] if n%2 else (s[n//2-1]+s[n//2])/2
        variance = sum((v-mean)**2 for v in s)/(n-1) if n>1 else 0
        std = math.sqrt(variance)
        q1 = s[max(0, n//4)]
        q3 = s[min(n-1, 3*n//4)]
        iqr = q3-q1
        return {"mean":mean,"median":median,"std":std,"q1":q1,"q3":q3,"iqr":iqr,
                "lower":q1-1.5*iqr,"upper":q3+1.5*iqr,"min":s[0],"max":s[-1]}

    prazos = df_analise["prazo_meses"].dropna().tolist()
    esforcos = df_analise["esforco"].dropna().tolist()
    st_prazos = stats(prazos)
    st_esforcos = stats(esforcos)

    # ── Correlações ──
    pearson, spearman = 0.0, 0.0
    reg_linear, reg_log = None, None

    if len(df_valido) >= 3:
        xs = df_valido["esforco"].tolist()
        ys = df_valido["prazo_meses"].tolist()
        pearson = calcular_pearson(xs, ys)
        spearman = calcular_spearman(xs, ys)
        if max(abs(pearson), abs(spearman)) >= 0.6:
            reg_linear = regressao_linear(xs, ys)
            reg_log = regressao_logaritmica(xs, ys)

    max_corr = max(abs(pearson), abs(spearman))
    corr_forte = max_corr >= 0.6

    # ── Seletor de Tipo de Regressão ──
    reg_type_sel = "auto"
    if corr_forte and reg_linear and reg_log:
        st.markdown("### Tipo de Regressão")
        auto_label = f"Automático ({'Linear' if abs(pearson) >= abs(spearman) else 'Logarítmica'})"
        opcoes_reg = {
            "auto": auto_label,
            "linear": f"Linear (Pearson: {pearson:.2f})",
            "logaritmica": f"Logarítmica (Spearman: {spearman:.2f})",
        }
        reg_cols = st.columns(3)
        for i, (key, label) in enumerate(opcoes_reg.items()):
            with reg_cols[i]:
                if st.button(label, key=f"reg_{key}", use_container_width=True,
                             type="primary" if st.session_state.get("reg_type_ap", "auto") == key else "secondary"):
                    st.session_state["reg_type_ap"] = key
                    st.rerun()
        reg_type_sel = st.session_state.get("reg_type_ap", "auto")

    # ── Determina regressão ativa ──
    reg = None
    reg_type_label = ""
    if corr_forte and reg_linear and reg_log:
        if reg_type_sel == "auto":
            if abs(pearson) >= abs(spearman):
                reg = reg_linear; reg_type_label = "Linear (Pearson dominante)"
            else:
                reg = reg_log; reg_type_label = "Logarítmica (Spearman dominante)"
        elif reg_type_sel == "linear":
            reg = reg_linear; reg_type_label = "Linear"
        else:
            reg = reg_log; reg_type_label = "Logarítmica"

    # ── Status correlação ──
    if corr_forte:
        st.success(f"✅ Correlação forte detectada! Máxima: {max_corr:.2f} (≥ 0,6) — Regressão **{reg_type_label}** será usada.")
    else:
        st.warning(f"⚠️ Correlação fraca ({max_corr:.2f} < 0,6). Serão usados os valores históricos mínimos e máximos.")

    # ── Gráfico dispersão ──
    if len(df_valido) >= 2:
        try:
            import plotly.graph_objects as go
            fig = go.Figure()
            fig.add_trace(go.Scatter(
                x=df_valido["esforco"].tolist(), y=df_valido["prazo_meses"].tolist(),
                mode="markers",
                marker=dict(size=10, color="#2563eb", opacity=0.7, line=dict(color="#1d4ed8", width=1)),
                text=df_valido.get("nome_edital", df_valido.index).tolist(),
                hovertemplate="<b>%{text}</b><br>Esforço: %{x}<br>Prazo: %{y:.1f} meses<extra></extra>",
                name="Projetos"
            ))
            title = f"Esforço vs Duração — {subtema_sel}"
            if reg and corr_forte:
                xs_s = sorted(df_valido["esforco"].tolist())
                ys_r = [reg["predict"](x) for x in xs_s]
                fig.add_trace(go.Scatter(
                    x=xs_s, y=ys_r, mode="lines",
                    line=dict(color="#ef4444", width=2),
                    name=f"Regressão {reg_type_label}",
                ))
                title = f"Regressão {reg_type_label}: {reg['eq']}"
            fig.update_layout(
                title=dict(text=title, font=dict(color="#ef4444", size=13)),
                xaxis_title="Esforço", yaxis_title="Duração (meses)",
                height=420, template="plotly_white",
                legend=dict(orientation="h", yanchor="bottom", y=1.02),
                margin=dict(t=60, b=40, l=40, r=20)
            )
            st.plotly_chart(fig, use_container_width=True)
        except ImportError:
            st.info("Instale plotly: pip install plotly")

    # ── Análise de Correlação ──
    st.markdown("### Análise de Correlação")
    cc1, cc2 = st.columns(2)
    def corr_badge(val):
        a = abs(val)
        if a >= 0.7: return "Forte"
        if a >= 0.6: return "Moderada"
        if a >= 0.3: return "Fraca"
        return "Muito fraca"
    def corr_dir(val):
        return "positiva" if val > 0 else "negativa" if val < 0 else "neutra"

    with cc1:
        em_uso_p = reg and reg.get("type") == "linear"
        badge_uso = ' <span style="background:#2563eb;color:#fff;padding:2px 8px;border-radius:999px;font-size:11px;font-weight:600;">✓ Em uso</span>' if em_uso_p else ''
        st.markdown(f"""
        <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:10px;padding:16px;">
            <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px;">
                <span style="color:#64748b;font-size:13px;">Pearson (Linear)</span>
                <span style="background:#e0f2fe;color:#0369a1;padding:2px 10px;border-radius:999px;font-size:11px;font-weight:600;">{corr_badge(pearson)}{badge_uso}</span>
            </div>
            <div style="font-size:28px;font-weight:800;color:#1e3a8a;">{pearson:.2f}</div>
            <div style="color:#94a3b8;font-size:12px;margin-top:4px;">Correlação {corr_dir(pearson)}</div>
        </div>
        """, unsafe_allow_html=True)

    with cc2:
        em_uso_s = reg and reg.get("type") == "logaritmica"
        badge_uso_s = ' <span style="background:#2563eb;color:#fff;padding:2px 8px;border-radius:999px;font-size:11px;font-weight:600;">✓ Em uso</span>' if em_uso_s else ''
        st.markdown(f"""
        <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:10px;padding:16px;">
            <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px;">
                <span style="color:#64748b;font-size:13px;">Spearman (Não-linear)</span>
                <span style="background:#e0f2fe;color:#0369a1;padding:2px 10px;border-radius:999px;font-size:11px;font-weight:600;">{corr_badge(spearman)}{badge_uso_s}</span>
            </div>
            <div style="font-size:28px;font-weight:800;color:#1e3a8a;">{spearman:.2f}</div>
            <div style="color:#94a3b8;font-size:12px;margin-top:4px;">Correlação {corr_dir(spearman)}</div>
        </div>
        """, unsafe_allow_html=True)

    if reg and corr_forte:
        st.markdown(f"""
        <div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:10px;padding:14px;margin-top:12px;">
            <div style="font-weight:700;color:#1e40af;margin-bottom:4px;">Equação da Regressão {reg_type_label}:</div>
            <div style="font-family:monospace;color:#1d4ed8;font-size:14px;">{reg['eq']}</div>
            <div style="font-weight:700;color:#1e40af;margin-top:8px;margin-bottom:4px;">Coeficiente de Determinação:</div>
            <div style="font-family:monospace;color:#1d4ed8;font-size:13px;">R² = {reg['r2']:.4f} &nbsp;&nbsp; 1 - R² = {1-reg['r2']:.4f}</div>
        </div>
        """, unsafe_allow_html=True)

    # ── Estatísticas ──
    st.markdown("### Estatísticas")
    sc1, sc2 = st.columns(2)
    def stat_table(data):
        return pd.DataFrame(list(data.items()), columns=["Métrica", "Valor"])
    with sc1:
        st.markdown("**Prazo (meses)**")
        if st_prazos:
            st.dataframe(stat_table({
                "Média": f"{st_prazos['mean']:.2f}", "Mediana": f"{st_prazos['median']:.2f}",
                "Desvio Padrão": f"{st_prazos['std']:.2f}", "Q1": f"{st_prazos['q1']:.2f}",
                "Q3": f"{st_prazos['q3']:.2f}", "IQR": f"{st_prazos['iqr']:.2f}",
                "Mínimo histórico": f"{st_prazos['min']:.2f}", "Máximo histórico": f"{st_prazos['max']:.2f}",
            }), hide_index=True, use_container_width=True)
    with sc2:
        st.markdown("**Esforço**")
        if st_esforcos:
            st.dataframe(stat_table({
                "Média": f"{st_esforcos['mean']:.2f}", "Mediana": f"{st_esforcos['median']:.2f}",
                "Desvio Padrão": f"{st_esforcos['std']:.2f}", "Q1": f"{st_esforcos['q1']:.2f}",
                "Q3": f"{st_esforcos['q3']:.2f}", "IQR": f"{st_esforcos['iqr']:.2f}",
                "Mínimo histórico": f"{st_esforcos['min']:.2f}", "Máximo histórico": f"{st_esforcos['max']:.2f}",
            }), hide_index=True, use_container_width=True)

    # ── Calculadora Kerzner ──
    st.markdown("### Calculadora de Prazos (Metodologia Kerzner)")
    if is_san:
        st.info("🌿 Tema SAN — Encerramento = 10% da execução.")

    esforco_input = 0.0
    if corr_forte and reg:
        unidade = ""
        if "unidade" in df_analise.columns:
            vals = df_analise["unidade"].dropna()
            if not vals.empty: unidade = vals.iloc[0]
        esforco_input = st.number_input(f"Esforço ({unidade})", min_value=0.0, step=0.1, key="ap_esforco")

    if st.button("Calcular Prazos", type="primary", key="ap_calcular"):
        hist_min = st_prazos.get("min", 0)
        hist_max = st_prazos.get("max", 0)
        if corr_forte and reg and esforco_input > 0:
            y_pred = reg["predict"](max(0.001, esforco_input))
            comp = max(0.1, 1 - reg["r2"])
            exec_min = max(0.5, y_pred * (1 - comp))
            exec_max = y_pred * (1 + comp)
            metodo = f"Regressão {reg_type_label}"
        else:
            exec_min = hist_min
            exec_max = hist_max
            metodo = "Valores históricos (correlação fraca)"

        total_min = exec_min / 0.4
        total_max = exec_max / 0.4
        plan_min = total_min * 0.5
        plan_max = total_max * 0.5
        enc_min = exec_min * 0.1 if is_san else max(1, total_min * 0.1)
        enc_max = exec_max * 0.1 if is_san else max(1, total_max * 0.1)

        kerzner_alerts = []
        if exec_min > hist_max:
            kerzner_alerts.append(f"Prazo mínimo de execução estimado ({exec_min:.1f}m) acima do máximo histórico ({hist_max:.1f}m)")
        if exec_max > hist_max:
            kerzner_alerts.append(f"Prazo máximo de execução estimado ({exec_max:.1f}m) acima do máximo histórico ({hist_max:.1f}m)")
        st.session_state["kerzner_alerts"] = kerzner_alerts
        st.session_state["kerzner_result"] = {
            "exec": (exec_min, exec_max), "plan": (plan_min, plan_max),
            "enc": (enc_min, enc_max), "total": (total_min, total_max),
            "metodo": metodo, "hist_exec_min": hist_min, "hist_exec_max": hist_max,
            "is_san": is_san, "reg_label": reg_type_label,
            "reg_eq": reg["eq"] if reg else "", "r2": reg["r2"] if reg else 0,
        }

    if "kerzner_result" in st.session_state:
        r = st.session_state["kerzner_result"]
        exec_min, exec_max = r["exec"]
        plan_min, plan_max = r["plan"]
        enc_min, enc_max = r["enc"]
        total_min, total_max = r["total"]
        hist_exec_min, hist_exec_max = r["hist_exec_min"], r["hist_exec_max"]
        enc_desc = "10% da execução" if r["is_san"] else "10% do total (mín. 1 mês)"

        st.markdown("#### Cronograma Completo do Projeto")
        fases = [
            ("🟡 Planejamento", "50% do projeto total", plan_min, plan_max, "#f59e0b"),
            ("🔵 Execução", "40% do projeto total", exec_min, exec_max, "#3b82f6"),
            (f"🟢 Encerramento", enc_desc, enc_min, enc_max, "#10b981"),
            ("🟣 Total", "Projeto completo", total_min, total_max, "#8b5cf6"),
        ]
        k1, k2, k3, k4 = st.columns(4)
        for col, (titulo, desc, vmin, vmax, cor) in zip([k1,k2,k3,k4], fases):
            with col:
                st.markdown(f"""
                <div style="border:1px solid #e2e8f0;border-left:4px solid {cor};border-radius:10px;padding:14px;background:#fff;">
                    <div style="font-weight:700;color:#0f172a;font-size:13px;margin-bottom:4px;">{titulo}</div>
                    <div style="color:#94a3b8;font-size:11px;margin-bottom:10px;">{desc}</div>
                    <div style="color:#64748b;font-size:12px;">Mínimo:</div>
                    <div style="font-size:22px;font-weight:800;color:{cor};">{vmin:.2f} <span style="font-size:13px;font-weight:400;">meses</span></div>
                    <div style="color:#64748b;font-size:12px;margin-top:6px;">Máximo:</div>
                    <div style="font-size:22px;font-weight:800;color:{cor};">{vmax:.2f} <span style="font-size:13px;font-weight:400;">meses</span></div>
                </div>
                """, unsafe_allow_html=True)

        # Alerta — só quando execução estimada > máximo histórico de execução
        alerts = []
        if exec_min > hist_exec_max:
            alerts.append(f"Prazo mínimo de execução estimado ({exec_min:.1f} m) está **acima** do máximo histórico observado ({hist_exec_max:.1f} m)")
        if exec_max > hist_exec_max:
            alerts.append(f"Prazo máximo de execução estimado ({exec_max:.1f} m) está **acima** do máximo histórico observado ({hist_exec_max:.1f} m)")
        if alerts:
            st.warning("⚠️ **Atenção: estimativa de execução fora do intervalo histórico**\n\n" + "\n\n".join(alerts) +
                      f"\n\nIntervalo histórico de execução: **{hist_exec_min:.1f}** a **{hist_exec_max:.1f}** meses. Use os valores com cautela.")

        # Cronograma visual
        st.markdown("#### Cronograma Visual (Metodologia Kerzner)")
        st.markdown("""
        <div style="display:flex;gap:16px;font-size:12px;color:#64748b;margin-bottom:12px;">
            <span><span style="display:inline-block;width:12px;height:12px;background:#f59e0b;border-radius:3px;margin-right:4px;"></span>Planejamento (50%)</span>
            <span><span style="display:inline-block;width:12px;height:12px;background:#3b82f6;border-radius:3px;margin-right:4px;"></span>Execução (40%)</span>
            <span><span style="display:inline-block;width:12px;height:12px;background:#10b981;border-radius:3px;margin-right:4px;"></span>Encerramento</span>
        </div>
        """, unsafe_allow_html=True)
        for label, (pv, ev, cv), total in [
            ("Cenário Mínimo", (plan_min, exec_min, enc_min), total_min),
            ("Cenário Máximo", (plan_max, exec_max, enc_max), total_max),
        ]:
            if total > 0:
                pp = pv/total*100; ep = ev/total*100; cp = cv/total*100
                st.markdown(f"**{label}** — {total:.2f} meses")
                st.markdown(f"""
                <div style="display:flex;height:28px;border-radius:8px;overflow:hidden;margin-bottom:4px;">
                    <div style="width:{pp:.1f}%;background:#f59e0b;"></div>
                    <div style="width:{ep:.1f}%;background:#3b82f6;"></div>
                    <div style="width:{cp:.1f}%;background:#10b981;"></div>
                </div>
                <div style="display:flex;justify-content:space-between;font-size:11px;color:#94a3b8;margin-bottom:14px;">
                    <span>0</span><span>{total:.2f} meses</span>
                </div>
                """, unsafe_allow_html=True)

        # Metodologia
        enc_met = "10% do tempo de execução (tema SAN)" if r["is_san"] else "10% do tempo total (mínimo 1 mês)"
        st.markdown(f"""
        <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:10px;padding:16px;margin-top:8px;">
            <div style="font-weight:700;color:#0f172a;margin-bottom:10px;">Metodologia Kerzner (2009)</div>
            <div style="font-size:13px;color:#475569;line-height:1.8;">
                • <b>Planejamento:</b> 50% do tempo total<br>
                • <b>Execução:</b> 40% do tempo total<br>
                • <b>Encerramento:</b> {enc_met}<br>
                • <b>Método de Cálculo:</b> {r['metodo']}
            </div>
        </div>
        """, unsafe_allow_html=True)

    # ── Classificação dos projetos ──
    st.markdown("### Classificação dos Projetos")

    def classificar(prazo, excluido):
        if excluido: return "Excluído"
        if not st_prazos: return "Normal"
        if prazo < st_prazos["lower"]: return "Outlier Baixo"
        if prazo >= st_prazos["upper"]: return "Outlier Alto"
        return "Normal"

    df_class = df_tema.copy()
    df_class["Classificação"] = df_class.apply(
        lambda row: classificar(row.get("prazo_meses", 0), row["_excluido"]), axis=1)

    n_norm = len(df_class[df_class["Classificação"] == "Normal"])
    n_low = len(df_class[df_class["Classificação"] == "Outlier Baixo"])
    n_high = len(df_class[df_class["Classificação"] == "Outlier Alto"])
    rc1, rc2, rc3, rc4 = st.columns(4)
    rc1.metric("Total", len(df_class))
    rc2.metric("Normal", n_norm)
    rc3.metric("Outlier Baixo", n_low)
    rc4.metric("Outlier Alto", n_high)

    st.markdown("**Marque projetos para excluir da análise:**")
    hc = st.columns([0.5, 3, 1.5, 1.5, 1, 1, 1.5])
    for col, label in zip(hc, ["", "Nome", "País", "Estado", "Esforço", "Prazo (m)", "Classificação"]):
        col.markdown(f"**{label}**")

    for _, row in df_class.iterrows():
        cols = st.columns([0.5, 3, 1.5, 1.5, 1, 1, 1.5])
        excluido = row["id"] in st.session_state[excluir_key]
        with cols[0]:
            novo = st.checkbox("", value=excluido, key=f"excl_{row['id']}_{excluir_key}", label_visibility="collapsed")
            if novo != excluido:
                if novo: st.session_state[excluir_key].add(row["id"])
                else: st.session_state[excluir_key].discard(row["id"])
                st.rerun()
        nome = str(row.get("nome_edital", "-") or "-")
        cols[1].write(f"~~{nome[:60]}~~" if excluido else nome[:60])
        cols[2].write(str(row.get("pais") or "-"))
        cols[3].write(str(row.get("estado") or "-"))
        cols[4].write(f"{row.get('esforco') or 0:.1f}")
        cols[5].write(f"{abs(row.get('prazo_meses') or 0):.1f}")
        cols[6].write(row["Classificação"])

    st.markdown("---")
    with st.container():
        kr_res = st.session_state.get("kerzner_result", {})
        export_data = {
            "tema": tema_sel if tema_sel != "Todos" else "",
            "subtema": subtema_sel if subtema_sel != "Todos" else tema_sel, "pais_sel": pais_sel, "estado_sel": estado_sel,
            "is_san": is_san, "pearson": pearson, "spearman": spearman,
            "corr_forte": corr_forte,
            "reg": {"type": reg.get("type",""),"eq": reg.get("eq",""),
                    "r2": reg.get("r2",0),"a": reg.get("a",0),"b": reg.get("b",0)} if reg else None,
            "reg_label": reg_type_label,
            "reg_type_display": f"Automático ({reg_type_label})" if reg_type_sel=="auto" else reg_type_label,
            "st_prazos": st_prazos, "st_esforcos": st_esforcos,
            "kerzner": {
                "plan_min": kr_res.get("plan",(0,0))[0],"plan_max": kr_res.get("plan",(0,0))[1],
                "exec_min": kr_res.get("exec",(0,0))[0],"exec_max": kr_res.get("exec",(0,0))[1],
                "enc_min":  kr_res.get("enc",(0,0))[0], "enc_max":  kr_res.get("enc",(0,0))[1],
                "total_min":kr_res.get("total",(0,0))[0],"total_max":kr_res.get("total",(0,0))[1],
            } if kr_res else {},
            "alerts": st.session_state.get("kerzner_alerts", []),
            "projetos": df_class.to_dict("records"),
        }
        try:
            xlsx_bytes = exportar_analise_excel(export_data)
            fname = f"Relatorio_Analise_{subtema_sel.replace(' ','_')}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
            st.download_button("📊 Exportar Relatório Excel", xlsx_bytes, file_name=fname,
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)
        except Exception as ex:
            st.error(f"Erro ao gerar Excel: {ex}")


# =========================================================
# MAIN
# =========================================================
def main():
    init_db()
    init_session()

    tema_visual = st.session_state.get("tema_visual", "Light")
    aplicar_estilo("dark" if tema_visual == "Dark" else "light")

    if not st.session_state.logado:
        tela_login()
        return

    menu_sidebar()

    if st.session_state.menu == "Base de Prazos":
        pagina_consulta()
    elif st.session_state.menu == "Análise de Prazos":
        pagina_analise_prazos()
    elif st.session_state.menu == "Projetos Concluídos":
        pagina_projetos_concluidos()
    elif st.session_state.menu == "Solicitações" or st.session_state.menu.startswith("Solicitações ("):
        pagina_solicitacoes()
    elif st.session_state.menu == "Base de dados":
        pagina_base()
    elif st.session_state.menu == "Minha conta":
        pagina_minha_conta()
    elif st.session_state.menu == "Usuários":
        pagina_usuarios()


if __name__ == "__main__":
    main()
